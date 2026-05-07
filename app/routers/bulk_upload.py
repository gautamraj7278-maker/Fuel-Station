from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO, StringIO
import csv
from typing import Any, Dict, List, Optional, Tuple
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app import models, schemas
from app.routers import sales as sales_router
from app.auth import get_password_hash
from app.database import get_db
from app.routers.auth import require_admin

router = APIRouter()


def _normalize_fuel_type(value: Any) -> str:
    return str(value or "").strip().lower()


def _get_category_names(db: Session) -> set[str]:
    rows = db.query(models.ProductCategory.name).all()
    return {str(r[0]).strip().lower() for r in rows if r and str(r[0]).strip()}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = s.replace(" ", "_")
    s = s.replace("-", "_")
    return s


def _to_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return int(value)
    s = str(value).strip()
    if s == "" or s.lower() in ("none", "null", "na"):
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s == "" or s.lower() in ("none", "null", "na"):
        return None
    try:
        return float(s)
    except Exception:
        return None


def _to_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    if s == "" or s.lower() in ("none", "null", "na"):
        return None
    try:
        return date.fromisoformat(s)
    except Exception:
        return None


def _to_time(value: Any) -> Optional[time]:
    if value is None:
        return None
    if isinstance(value, time):
        return value
    s = str(value).strip()
    if s == "" or s.lower() in ("none", "null", "na"):
        return None
    try:
        return time.fromisoformat(s)
    except Exception:
        return None


def _to_bool(value: Any) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("true", "1", "yes", "y"):
        return True
    if s in ("false", "0", "no", "n"):
        return False
    return None


def _validate_required(row: Dict[str, Any], required: List[str]) -> List[str]:
    errors = []
    for f in required:
        v = row.get(f)
        if v is None or str(v).strip() == "":
            errors.append(f"{f} is required")
    return errors


def _read_csv_bytes(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    f = StringIO(text)
    reader = csv.DictReader(f)
    raw_headers = list(reader.fieldnames or [])
    headers = [_normalize_header(h) for h in raw_headers]

    rows: List[Dict[str, Any]] = []
    for row in reader:
        normalized: Dict[str, Any] = {}
        for raw_key, value in row.items():
            normalized[_normalize_header(raw_key)] = value
        rows.append(normalized)

    # Prefer normalized headers list.
    headers = [h for h in headers if h]
    if not headers and rows:
        headers = sorted({k for r in rows for k in r.keys()})

    return headers, rows


def _read_xlsx_bytes(content: bytes, *, sheet_name_hint: Optional[str] = None) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = load_workbook(filename=BytesIO(content), data_only=True)

    ws = None
    if sheet_name_hint:
        for name in wb.sheetnames:
            if name.strip().lower() == sheet_name_hint.strip().lower():
                ws = wb[name]
                break
    if ws is None:
        ws = wb.active

    header_row = None
    for r in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if not r:
            continue
        if any(v is not None and str(v).strip() != "" for v in r):
            header_row = r
            break

    if not header_row:
        return [], []

    headers = [_normalize_header(h) for h in header_row]

    rows: List[Dict[str, Any]] = []
    # Find the header row index
    header_index = None
    for idx, r in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if tuple(r) == tuple(header_row):
            header_index = idx
            break
    if header_index is None:
        header_index = 1

    for r in ws.iter_rows(min_row=header_index + 1, values_only=True):
        if not r or not any(v is not None and str(v).strip() != "" for v in r):
            continue
        row_obj: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            row_obj[h] = r[i] if i < len(r) else None
        rows.append(row_obj)

    headers = [h for h in headers if h]
    if not headers and rows:
        headers = sorted({k for rr in rows for k in rr.keys()})

    return headers, rows


def _generate_transaction_id() -> str:
    return f"TXN{datetime.utcnow().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"


SALES_TEMPLATE_COLUMNS = [
    "business_date",  # optional (YYYY-MM-DD). Leave blank to auto-calc.
    "shift",  # optional (A/B/C). Leave blank to auto-calc.
    "transaction_type",  # sale/testing
    "operator_employee_id",  # required
    "dispenser_id",  # required
    "nozzle_id",  # required if no meter_id
    "meter_id",  # required if using meters
    "closing_meter_reading",  # required if meter_id
    "quantity",  # required if no meter_id (liters)
    "testing_quantity",  # optional (liters)
    "deposit_cash",  # optional
    "deposit_online",  # optional
    "remarks",  # optional
]

SALES_PREVIEW_COLUMNS = SALES_TEMPLATE_COLUMNS + [
    "fuel_type",
    "product_id",
    "opening_meter_reading",
    "dispensed_quantity",
    "price_per_liter",
    "total_amount",
    "total_deposit",
]


TANKER_RECEIPTS_TEMPLATE_RECEIPTS_COLUMNS = [
    "receipt_id",  # optional (for updating draft receipts)
    "receipt_key",  # required (used to join with Lines sheet)
    "receipt_date",  # YYYY-MM-DD
    "tanker_no",
    "transporter_name",
    "driver_name",
    "invoice_no",
    "remarks",
]


TANKER_RECEIPTS_TEMPLATE_LINES_COLUMNS = [
    "receipt_key",  # required
    "tank_id",
    "before_dips_mm",
    "after_dips_mm",
    "dips_invoice_mm",  # optional (per product/compartment)
    "dips_site_mm",  # optional (per product/compartment)
    "quantity_invoice_litres",  # optional (per product/compartment)
    "density_invoice",  # optional (per product/compartment)
    "density_site",  # optional (per product/compartment)
    "temperature_c",  # optional (per product/compartment)
    "compartment_remarks",  # optional (per product/compartment)
    "remarks",
]


TANK_CALIBRATION_POINTS_TEMPLATE_COLUMNS = [
    "calibration_point_id",  # optional (for updates)
    "tank_id",
    "dips_mm",
    "volume_in_litres",
]


FUEL_INVENTORY_TEMPLATE_COLUMNS = [
    "fuel_type",  # product category
    "action",  # set/add/subtract
    "quantity",  # liters
    "price_per_liter",  # optional
    "reorder_level",  # optional
    "notes",  # optional
]

FUEL_INVENTORY_PREVIEW_COLUMNS = FUEL_INVENTORY_TEMPLATE_COLUMNS + [
    "current_stock",
    "new_stock",
]


TANK_TRANSFERS_TEMPLATE_COLUMNS = [
    "transfer_id",  # optional (for updates)
    "from_tank_id",
    "to_tank_id",
    "product_id",
    "volume",
    "transfer_type",  # manual/testing_to_buffer/buffer_to_main
    "user_id",  # optional (defaults to current user)
]


DISPENSER_SHIFT_ASSIGNMENTS_TEMPLATE_COLUMNS = [
    "assignment_id",  # optional (for updates)
    "business_date",  # YYYY-MM-DD
    "shift",  # A/B/C
    "dispenser_id",
    "operator_id",
]


DAILY_CLOSES_TEMPLATE_COLUMNS = [
    "daily_close_id",  # optional (for updates)
    "business_date",  # YYYY-MM-DD
    "user_id",
    "opening_cash",
    "closing_cash",
    "notes",
]


USERS_TEMPLATE_COLUMNS = [
    "user_id",  # optional (for updates)
    "username",
    "email",
    "password",  # required for inserts
    "full_name",
    "role",  # admin/manager/operator
    "is_active",
]


PRODUCTS_TEMPLATE_COLUMNS = [
    "product_name",
    "fuel_type",  # product category
    "is_active",  # true/false
]


PRODUCT_PRICES_TEMPLATE_COLUMNS = [
    "product_id",  # optional if product_name provided
    "product_name",  # optional if product_id provided
    "effective_date",  # YYYY-MM-DD
    "price_per_liter",
    "remarks",
]


DISPENSERS_TEMPLATE_COLUMNS = [
    "dispenser_number",
    "is_active",  # true/false
]


NOZZLES_TEMPLATE_COLUMNS = [
    "dispenser_id",  # optional if dispenser_number provided
    "dispenser_number",  # optional if dispenser_id provided
    "nozzle_number",
    "product_id",  # optional if product_name provided
    "product_name",  # optional if product_id provided
    "tank_id",  # optional if tank_name provided
    "tank_name",  # optional if tank_id provided
    "is_active",
]


METERS_TEMPLATE_COLUMNS = [
    "nozzle_id",  # required (or nozzle_number+dispenser_number)
    "dispenser_number",  # optional lookup
    "nozzle_number",  # optional lookup
    "meter_name",
    "max_value",
    "last_reading",
    "is_active",
]


TANKS_TEMPLATE_COLUMNS = [
    "tank_name",
    "product_id",  # optional if product_name provided
    "product_name",  # optional if product_id provided
    "capacity",
    "current_volume",
    "is_buffer",
    "calibration_date",  # YYYY-MM-DD
    "calibration_due_date",  # YYYY-MM-DD
    "remarks",
]


EMPLOYEES_TEMPLATE_COLUMNS = [
    "employee_name",
    "dob",  # YYYY-MM-DD
    "address",
    "contact_no",
    "id_no",
    "designation_id",  # optional if designation_name provided
    "designation_name",
    "is_active",
]


DESIGNATIONS_TEMPLATE_COLUMNS = [
    "name",
    "is_active",
]


SHIFTS_TEMPLATE_COLUMNS = [
    "shift",  # A/B/C
    "start_time",  # HH:MM:SS
    "end_time",  # HH:MM:SS
    "is_active",
    "remarks",
]


TANK_DIPS_TEMPLATE_COLUMNS = [
    "tank_id",  # optional if tank_name provided
    "tank_name",
    "business_date",  # YYYY-MM-DD
    "dip_type",  # opening/closing
    "dips_mm",
    "computed_volume_litres",
    "manual_volume_litres",
    "is_auto",
]


def _interpolate_volume(points: List[models.TankCalibrationPoint], dips_mm: float) -> float:
    if dips_mm < 0:
        raise HTTPException(status_code=400, detail="Dips must be >= 0")

    pts = sorted(points, key=lambda p: float(p.dips_mm))
    if not pts:
        raise HTTPException(status_code=400, detail="No calibration points found for tank")

    if dips_mm < float(pts[0].dips_mm) or dips_mm > float(pts[-1].dips_mm):
        raise HTTPException(
            status_code=400,
            detail=f"Dip {dips_mm}mm is outside calibration range ({pts[0].dips_mm}mm - {pts[-1].dips_mm}mm)",
        )

    for p in pts:
        if float(p.dips_mm) == float(dips_mm):
            return float(p.volume_in_litres)

    lower = None
    upper = None
    for p in pts:
        if float(p.dips_mm) < dips_mm:
            lower = p
        elif float(p.dips_mm) > dips_mm:
            upper = p
            break

    if lower is None or upper is None:
        raise HTTPException(status_code=400, detail="Unable to interpolate dip")

    x0 = float(lower.dips_mm)
    y0 = float(lower.volume_in_litres)
    x1 = float(upper.dips_mm)
    y1 = float(upper.volume_in_litres)
    if x1 == x0:
        return float(y0)
    ratio = (dips_mm - x0) / (x1 - x0)
    return float(y0 + ratio * (y1 - y0))


def _compute_tanker_line(db: Session, *, tank: models.Tank, before_dips_mm: float, after_dips_mm: float):
    points = (
        db.query(models.TankCalibrationPoint)
        .filter(models.TankCalibrationPoint.tank_id == tank.id)
        .order_by(models.TankCalibrationPoint.dips_mm.asc())
        .all()
    )
    before_volume = _interpolate_volume(points, before_dips_mm)
    after_volume = _interpolate_volume(points, after_dips_mm)
    received = after_volume - before_volume
    if received < 0:
        raise HTTPException(status_code=400, detail="After dip volume is less than before dip volume")
    return before_volume, after_volume, received


@router.get("/tanker-receipts/template")
def download_tanker_receipts_template(
    current_user: models.User = Depends(require_admin),
):
    _ = current_user

    wb = Workbook()
    ws_r = wb.active
    ws_r.title = "Receipts"
    ws_r.append(TANKER_RECEIPTS_TEMPLATE_RECEIPTS_COLUMNS)
    ws_r.append([
        "",  # receipt_id
        "R1",
        date.today().isoformat(),
        "TN-001",
        "Transporter",
        "Driver",
        "INV-001",
        "Backfill example",
    ])

    ws_l = wb.create_sheet("Lines")
    ws_l.append(TANKER_RECEIPTS_TEMPLATE_LINES_COLUMNS)
    ws_l.append(["R1", 1, 100, 120, 98, 102, 5000, 0.83, 0.82, 28, "Compartment notes", "Tank line example"])

    notes = wb.create_sheet("Notes")
    notes.append(["How to use"])
    notes.append(["- Fill Receipts + Lines, then upload in Tanker Receipts → Bulk Upload."])
    notes.append(["- receipt_key must match between both sheets."])
    notes.append(["- receipt_id is optional and only used to update an existing DRAFT receipt."])
    notes.append(["- Compartment fields (dips/density/temperature) are set per product and shared across lines."])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tanker_receipts_bulk_template.xlsx"},
    )


@router.post("/tanker-receipts/preview")
async def preview_tanker_receipts_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user

    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .xlsx")

    wb = load_workbook(filename=BytesIO(content), data_only=True)
    receipts_ws = None
    lines_ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == "receipts":
            receipts_ws = wb[name]
        if name.strip().lower() == "lines":
            lines_ws = wb[name]
    if receipts_ws is None or lines_ws is None:
        raise HTTPException(status_code=400, detail="XLSX must contain sheets named 'Receipts' and 'Lines'")

    def read_sheet(ws) -> Tuple[List[str], List[Dict[str, Any]]]:
        headers = None
        header_row_idx = None
        for idx, r in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if r and any(v is not None and str(v).strip() != "" for v in r):
                headers = [_normalize_header(v) for v in r]
                header_row_idx = idx
                break
        if not headers or header_row_idx is None:
            return [], []
        out_rows = []
        for ridx, r in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            if not r or not any(v is not None and str(v).strip() != "" for v in r):
                continue
            obj = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                obj[h] = r[i] if i < len(r) else None
            obj["_row"] = ridx
            out_rows.append(obj)
        return [h for h in headers if h], out_rows

    _, receipt_rows = read_sheet(receipts_ws)
    _, line_rows = read_sheet(lines_ws)

    if len(receipt_rows) > 2000 or len(line_rows) > 5000:
        raise HTTPException(status_code=400, detail="Too many rows. Please split the upload into smaller files.")

    errors = []
    receipts_by_key: Dict[str, Dict[str, Any]] = {}
    for r in receipt_rows:
        key = str(r.get("receipt_key") or "").strip()
        if not key:
            errors.append({"sheet": "Receipts", "row": int(r.get("_row") or 0), "message": "receipt_key is required"})
            continue
        if key in receipts_by_key:
            errors.append({"sheet": "Receipts", "row": int(r.get("_row") or 0), "message": f"Duplicate receipt_key '{key}'"})
            continue
        receipt_date = _to_date(r.get("receipt_date"))
        if receipt_date is None:
            errors.append({"sheet": "Receipts", "row": int(r.get("_row") or 0), "message": "receipt_date is required (YYYY-MM-DD)"})
            continue
        tanker_no = str(r.get("tanker_no") or "").strip()
        if not tanker_no:
            errors.append({"sheet": "Receipts", "row": int(r.get("_row") or 0), "message": "tanker_no is required"})
            continue

        receipts_by_key[key] = {
            "receipt_id": _to_int(r.get("receipt_id")),
            "receipt_key": key,
            "receipt_date": receipt_date.isoformat(),
            "tanker_no": tanker_no,
            "transporter_name": (str(r.get("transporter_name") or "").strip() or None),
            "driver_name": (str(r.get("driver_name") or "").strip() or None),
            "invoice_no": (str(r.get("invoice_no") or "").strip() or None),
            "remarks": (str(r.get("remarks") or "").strip() or None),
            "lines": [],
            "_row": int(r.get("_row") or 0),
        }

    for l in line_rows:
        key = str(l.get("receipt_key") or "").strip()
        if not key:
            errors.append({"sheet": "Lines", "row": int(l.get("_row") or 0), "message": "receipt_key is required"})
            continue
        parent = receipts_by_key.get(key)
        if parent is None:
            errors.append({"sheet": "Lines", "row": int(l.get("_row") or 0), "message": f"Unknown receipt_key '{key}'"})
            continue

        tank_id = _to_int(l.get("tank_id"))
        before_mm = _to_float(l.get("before_dips_mm"))
        after_mm = _to_float(l.get("after_dips_mm"))
        dips_invoice_mm = _to_float(l.get("dips_invoice_mm"))
        dips_site_mm = _to_float(l.get("dips_site_mm"))
        quantity_invoice_litres = _to_float(l.get("quantity_invoice_litres"))
        density_invoice = _to_float(l.get("density_invoice"))
        density_site = _to_float(l.get("density_site"))
        temperature_c = _to_float(l.get("temperature_c"))
        compartment_remarks = str(l.get("compartment_remarks") or "").strip() or None
        if tank_id is None:
            errors.append({"sheet": "Lines", "row": int(l.get("_row") or 0), "message": "tank_id is required"})
            continue
        if before_mm is None or after_mm is None:
            errors.append({"sheet": "Lines", "row": int(l.get("_row") or 0), "message": "before_dips_mm and after_dips_mm are required"})
            continue

        tank = db.query(models.Tank).filter(models.Tank.id == tank_id).first()
        if not tank:
            errors.append({"sheet": "Lines", "row": int(l.get("_row") or 0), "message": f"Tank not found: {tank_id}"})
            continue

        try:
            before_vol, after_vol, received = _compute_tanker_line(db, tank=tank, before_dips_mm=float(before_mm), after_dips_mm=float(after_mm))
        except HTTPException as exc:
            errors.append({"sheet": "Lines", "row": int(l.get("_row") or 0), "message": str(exc.detail)})
            continue

        parent["lines"].append(
            {
                "tank_id": tank_id,
                "product_id": int(tank.product_id),
                "before_dips_mm": float(before_mm),
                "after_dips_mm": float(after_mm),
                "before_volume_litres": float(before_vol),
                "after_volume_litres": float(after_vol),
                "received_volume_litres": float(received),
                "dips_invoice_mm": dips_invoice_mm,
                "dips_site_mm": dips_site_mm,
                "quantity_invoice_litres": quantity_invoice_litres,
                "density_invoice": density_invoice,
                "density_site": density_site,
                "temperature_c": temperature_c,
                "compartment_remarks": compartment_remarks,
                "remarks": (str(l.get("remarks") or "").strip() or None),
                "_row": int(l.get("_row") or 0),
            }
        )

    receipts = list(receipts_by_key.values())
    for r in receipts:
        if not r.get("lines"):
            errors.append({"sheet": "Receipts", "row": int(r.get("_row") or 0), "message": f"No lines found for receipt_key '{r.get('receipt_key')}'"})

    valid_receipts = [r for r in receipts if r.get("lines")]
    return {
        "total_receipts": len(receipts),
        "valid_receipts": len(valid_receipts),
        "total_lines": len(line_rows),
        "errors": errors,
        "receipts": valid_receipts,
    }


@router.post("/tanker-receipts/commit")
def commit_tanker_receipts_upload(
    payload: Dict[str, Any],
    confirm: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    receipts = payload.get("receipts") or []
    if not isinstance(receipts, list) or not receipts:
        raise HTTPException(status_code=400, detail="No receipts to commit")

    inserted = 0
    updated = 0

    for r in receipts:
        receipt_id = _to_int(r.get("receipt_id"))
        receipt_date = _to_date(r.get("receipt_date"))
        tanker_no = str(r.get("tanker_no") or "").strip()
        if receipt_date is None or not tanker_no:
            raise HTTPException(status_code=400, detail="Invalid receipt payload")

        lines = r.get("lines") or []
        if not isinstance(lines, list) or not lines:
            raise HTTPException(status_code=400, detail="Receipt has no lines")

        db_receipt = None
        if receipt_id is not None:
            db_receipt = db.query(models.TankerReceipt).filter(models.TankerReceipt.id == receipt_id).first()
            if db_receipt and db_receipt.status != models.TankerReceiptStatus.DRAFT:
                raise HTTPException(status_code=400, detail=f"Receipt {receipt_id} is not draft; cannot bulk update")

        if db_receipt is None:
            db_receipt = models.TankerReceipt(
                receipt_date=receipt_date,
                tanker_no=tanker_no,
                transporter_name=r.get("transporter_name"),
                driver_name=r.get("driver_name"),
                invoice_no=r.get("invoice_no"),
                remarks=r.get("remarks"),
                status=models.TankerReceiptStatus.DRAFT,
                created_by_user_id=current_user.id,
            )
            db.add(db_receipt)
            db.flush()
            inserted += 1
        else:
            db_receipt.receipt_date = receipt_date
            db_receipt.tanker_no = tanker_no
            db_receipt.transporter_name = r.get("transporter_name")
            db_receipt.driver_name = r.get("driver_name")
            db_receipt.invoice_no = r.get("invoice_no")
            db_receipt.remarks = r.get("remarks")
            db_receipt.compartments.clear()
            db_receipt.lines.clear()
            db.flush()
            updated += 1

        compartments_by_key: Dict[Tuple[Any, ...], models.TankerReceiptCompartment] = {}
        compartments_by_product: Dict[int, List[models.TankerReceiptCompartment]] = {}
        has_compartment_data = False
        for l in lines:
            pid = _to_int(l.get("product_id"))
            if pid is None:
                continue
            fields = (
                int(pid),
                _to_float(l.get("dips_invoice_mm")),
                _to_float(l.get("dips_site_mm")),
                _to_float(l.get("quantity_invoice_litres")),
                _to_float(l.get("density_invoice")),
                _to_float(l.get("density_site")),
                _to_float(l.get("temperature_c")),
                (str(l.get("compartment_remarks") or "").strip() or None),
            )
            if any(v is not None for v in fields[1:]):
                has_compartment_data = True
                comp = compartments_by_key.get(fields)
                if comp is None:
                    comp = models.TankerReceiptCompartment(
                        receipt_id=db_receipt.id,
                        product_id=int(pid),
                        dips_invoice_mm=fields[1],
                        dips_site_mm=fields[2],
                        quantity_invoice_litres=fields[3],
                        density_invoice=fields[4],
                        density_site=fields[5],
                        temperature_c=fields[6],
                        remarks=fields[7],
                    )
                    db.add(comp)
                    db.flush()
                    compartments_by_key[fields] = comp
                    compartments_by_product.setdefault(int(pid), []).append(comp)

        for l in lines:
            tank_id = _to_int(l.get("tank_id"))
            before_mm = _to_float(l.get("before_dips_mm"))
            after_mm = _to_float(l.get("after_dips_mm"))
            if tank_id is None or before_mm is None or after_mm is None:
                raise HTTPException(status_code=400, detail="Invalid tanker line")

            tank = db.query(models.Tank).filter(models.Tank.id == tank_id).first()
            if not tank:
                raise HTTPException(status_code=404, detail=f"Tank not found: {tank_id}")

            before_vol, after_vol, received = _compute_tanker_line(db, tank=tank, before_dips_mm=float(before_mm), after_dips_mm=float(after_mm))
            pid = int(tank.product_id)
            comp = None
            if has_compartment_data:
                comp_key = (
                    int(pid),
                    _to_float(l.get("dips_invoice_mm")),
                    _to_float(l.get("dips_site_mm")),
                    _to_float(l.get("quantity_invoice_litres")),
                    _to_float(l.get("density_invoice")),
                    _to_float(l.get("density_site")),
                    _to_float(l.get("temperature_c")),
                    (str(l.get("compartment_remarks") or "").strip() or None),
                )
                if any(v is not None for v in comp_key[1:]):
                    comp = compartments_by_key.get(comp_key)
                else:
                    comp_list = compartments_by_product.get(pid) or []
                    comp = comp_list[0] if comp_list else None
                if comp is None:
                    raise HTTPException(status_code=400, detail="Tank product does not match any receipt compartment")

            db_receipt.lines.append(
                models.TankerReceiptLine(
                    compartment_id=(comp.id if comp is not None else None),
                    tank_id=tank.id,
                    product_id=pid,
                    before_dips_mm=float(before_mm),
                    after_dips_mm=float(after_mm),
                    before_volume_litres=float(before_vol),
                    after_volume_litres=float(after_vol),
                    received_volume_litres=float(received),
                    remarks=(str(l.get("remarks") or "").strip() or None),
                )
            )

        if confirm:
            # Apply stock updates (same as confirm endpoint)
            deltas_by_fuel_type: Dict[str, float] = {}
            for line in db_receipt.lines:
                tank = db.query(models.Tank).filter(models.Tank.id == line.tank_id).first()
                if not tank:
                    raise HTTPException(status_code=404, detail=f"Tank not found: {line.tank_id}")
                if bool(getattr(tank, "is_buffer", False)):
                    raise HTTPException(status_code=400, detail=f"Cannot confirm into buffer tank '{tank.tank_name}'")
                product = db.query(models.Product).filter(models.Product.id == tank.product_id).first()
                if not product:
                    raise HTTPException(status_code=400, detail="Tank product is missing")
                delta = float(line.received_volume_litres or 0)
                prev = float(tank.current_volume or 0)
                tank.current_volume = prev + delta
                deltas_by_fuel_type[product.fuel_type] = float(deltas_by_fuel_type.get(product.fuel_type, 0.0) + delta)
                db.add(
                    models.TankStockLog(
                        tank_id=tank.id,
                        action="tanker_confirm",
                        quantity=delta,
                        previous_volume=prev,
                        new_volume=float(tank.current_volume),
                        notes=f"Bulk-confirm tanker receipt {db_receipt.id}",
                        related_receipt_id=db_receipt.id,
                        created_by_user_id=current_user.id,
                    )
                )

            for fuel_type, delta in deltas_by_fuel_type.items():
                inv = db.query(models.FuelInventory).filter(models.FuelInventory.fuel_type == fuel_type).first()
                if inv is None:
                    inv = models.FuelInventory(
                        fuel_type=fuel_type,
                        current_stock=0.0,
                        price_per_liter=0.01,
                        reorder_level=0.0,
                    )
                    db.add(inv)
                    db.flush()
                previous_stock = float(inv.current_stock or 0.0)
                inv.current_stock = previous_stock + float(delta)
                db.add(
                    models.InventoryLog(
                        fuel_type=fuel_type,
                        action="tanker_receipt",
                        quantity=float(delta),
                        previous_stock=previous_stock,
                        new_stock=float(inv.current_stock),
                        notes=f"Bulk-confirm tanker receipt {db_receipt.id}",
                    )
                )

            db_receipt.status = models.TankerReceiptStatus.CONFIRMED
            db_receipt.confirmed_at = datetime.utcnow()
            db_receipt.confirmed_by_user_id = current_user.id

    db.commit()
    return {"inserted": inserted, "updated": updated}


def _validate_and_normalize_sale_row(row: Dict[str, Any], *, row_number: int) -> Tuple[Optional[Dict[str, Any]], List[str]]:
    errors: List[str] = []
    business_date = _to_date(row.get("business_date"))

    shift_raw = str(row.get("shift") or "").strip().upper() or None
    if shift_raw and shift_raw not in ("A", "B", "C"):
        errors.append("shift must be A, B, or C")

    tx_raw = str(row.get("transaction_type") or "sale").strip().lower()
    if tx_raw not in ("sale", "testing"):
        errors.append("transaction_type must be 'sale' or 'testing'")

    operator_employee_id = _to_int(row.get("operator_employee_id"))
    if operator_employee_id is None:
        errors.append("operator_employee_id is required")

    dispenser_id = _to_int(row.get("dispenser_id"))
    if dispenser_id is None:
        errors.append("dispenser_id is required")

    nozzle_id = _to_int(row.get("nozzle_id"))
    meter_id = _to_int(row.get("meter_id"))
    if meter_id is None and nozzle_id is None:
        errors.append("Either meter_id or nozzle_id is required")

    closing_meter_reading = _to_float(row.get("closing_meter_reading"))
    quantity = _to_float(row.get("quantity"))
    testing_quantity = _to_float(row.get("testing_quantity"))

    if meter_id is not None and closing_meter_reading is None:
        errors.append("closing_meter_reading is required when meter_id is provided")
    if meter_id is None and (quantity is None or quantity <= 0):
        errors.append("quantity is required and must be > 0 when meter_id is not provided")
    if testing_quantity is not None and testing_quantity < 0:
        errors.append("testing_quantity must be >= 0")

    deposit_cash = _to_float(row.get("deposit_cash"))
    deposit_online = _to_float(row.get("deposit_online"))
    if deposit_cash is not None and deposit_cash < 0:
        errors.append("deposit_cash must be >= 0")
    if deposit_online is not None and deposit_online < 0:
        errors.append("deposit_online must be >= 0")

    remarks = row.get("remarks")
    if remarks is not None:
        remarks = str(remarks).strip() or None

    if errors:
        return None, errors

    normalized = {
        "business_date": business_date.isoformat() if business_date else None,
        "shift": shift_raw,
        "transaction_type": tx_raw,
        "operator_employee_id": operator_employee_id,
        "dispenser_id": dispenser_id,
        "nozzle_id": nozzle_id,
        "meter_id": meter_id,
        "closing_meter_reading": closing_meter_reading,
        "quantity": float(quantity) if quantity is not None else None,
        "testing_quantity": float(testing_quantity) if testing_quantity is not None else None,
        "deposit_cash": float(deposit_cash or 0.0),
        "deposit_online": float(deposit_online or 0.0),
        "remarks": remarks,
        "_row": row_number,
    }

    return normalized, []


@router.get("/sales/template")
def download_sales_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user

    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(SALES_TEMPLATE_COLUMNS)
        writer.writerow(
            [
                date.today().isoformat(),
                "A",
                "sale",
                "1",
                "1",
                "1",
                "",  # meter_id
                "1500",  # closing_meter_reading
                "",  # quantity
                "0",  # testing_quantity
                "0",
                "0",
                "Manual entry example row",
            ]
        )
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=sales_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(SALES_TEMPLATE_COLUMNS)
    ws.append(
        [
            date.today().isoformat(),
            "A",
            "sale",
            1,
            1,
            1,
            "",  # meter_id
            "1500",  # closing_meter_reading
            "",  # quantity
            0,  # testing_quantity
            0,
            0,
            "Manual entry example row",
        ]
    )

    notes = wb.create_sheet("Notes")
    notes.append(["How to use"])
    notes.append(["- Fill the Sales sheet, then upload the file in Sales → Bulk Upload."])
    notes.append(["- Use the same fields as the manual Sales entry screen."])
    notes.append(["- quantity is required only when meter_id is blank (sales quantity)."])
    notes.append(["- testing_quantity is optional; it is subtracted from dispensed quantity."])
    notes.append(["- meter_id requires closing_meter_reading."])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sales_bulk_template.xlsx"},
    )


@router.post("/sales/preview", response_model=schemas.SalesBulkPreviewResponse)
async def preview_sales_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user

    filename = (file.filename or "").lower()
    content = await file.read()

    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        columns, raw_rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        columns, raw_rows = _read_xlsx_bytes(content, sheet_name_hint="Sales")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    max_rows = 5000
    if len(raw_rows) > max_rows:
        raise HTTPException(status_code=400, detail=f"Too many rows ({len(raw_rows)}). Please upload {max_rows} rows or fewer per file.")

    errors: List[schemas.BulkRowError] = []
    normalized_rows: List[Dict[str, Any]] = []

    for idx, row in enumerate(raw_rows, start=2):
        normalized, row_errors = _validate_and_normalize_sale_row(row, row_number=idx)
        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue
        assert normalized is not None
        try:
            sale_payload = schemas.SaleCreate(
                dispenser_id=int(normalized["dispenser_id"]),
                nozzle_id=normalized.get("nozzle_id"),
                meter_id=normalized.get("meter_id"),
                quantity=normalized.get("quantity"),
                testing_quantity=normalized.get("testing_quantity"),
                closing_meter_reading=normalized.get("closing_meter_reading"),
                business_date=_to_date(normalized.get("business_date")),
                transaction_type=models.TransactionType(normalized["transaction_type"]),
                shift=models.ShiftCode(normalized["shift"]) if normalized.get("shift") else models.ShiftCode.A,
                operator_employee_id=normalized.get("operator_employee_id"),
                deposit_cash=normalized.get("deposit_cash"),
                deposit_online=normalized.get("deposit_online"),
                remarks=normalized.get("remarks"),
            )
            preview = sales_router.compute_sale_preview(sale_payload, db=db, current_user=current_user)
            normalized_rows.append({**normalized, **preview})
        except HTTPException as e:
            errors.append(schemas.BulkRowError(row=idx, message=str(e.detail)))

    return schemas.SalesBulkPreviewResponse(
        columns=SALES_PREVIEW_COLUMNS,
        total_rows=len(raw_rows),
        valid_rows=len(normalized_rows),
        errors=errors,
        rows=normalized_rows,
    )


@router.post("/sales/commit", response_model=schemas.SalesBulkCommitResponse)
def commit_sales_upload(
    payload: schemas.SalesBulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    """Insert sales rows from raw-entry bulk upload."""
    inserted = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        normalized, errors = _validate_and_normalize_sale_row(row, row_number=int(row.get("_row") or i))
        if errors or normalized is None:
            for msg in errors or ["Invalid row"]:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message=msg))
            continue

        try:
            sale_payload = schemas.SaleCreate(
                dispenser_id=int(normalized["dispenser_id"]),
                nozzle_id=normalized.get("nozzle_id"),
                meter_id=normalized.get("meter_id"),
                quantity=normalized.get("quantity"),
                testing_quantity=normalized.get("testing_quantity"),
                closing_meter_reading=normalized.get("closing_meter_reading"),
                business_date=_to_date(normalized.get("business_date")),
                transaction_type=models.TransactionType(normalized["transaction_type"]),
                shift=models.ShiftCode(normalized["shift"]) if normalized.get("shift") else models.ShiftCode.A,
                operator_employee_id=normalized.get("operator_employee_id"),
                deposit_cash=normalized.get("deposit_cash"),
                deposit_online=normalized.get("deposit_online"),
                remarks=normalized.get("remarks"),
            )
            sales_router.create_sale(sale_payload, db=db, current_user=current_user)
            inserted += 1
        except HTTPException as e:
            row_errors.append(schemas.BulkRowError(row=int(normalized.get("_row") or i), message=str(e.detail)))

    if row_errors:
        # Don’t partially commit unless caller explicitly opts in.
        if not payload.allow_partial:
            db.rollback()
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "Validation failed",
                    "errors": [e.model_dump() for e in row_errors[:200]],
                },
            )

    db.commit()

    return schemas.SalesBulkCommitResponse(inserted=inserted, updated=0, failed=len(row_errors))


# -----------------------------
# Customers Bulk Upload
# -----------------------------

CUSTOMERS_TEMPLATE_COLUMNS = [
    "customer_id",  # optional (for updates)
    "name",  # required
    "phone",  # required (unique)
    "email",  # optional (unique)
    "vehicle_number",  # optional
]


def _normalize_phone(value: Any) -> str:
    s = str(value or "").strip()
    out: List[str] = []
    for ch in s:
        if ch.isdigit() or ch == "+":
            out.append(ch)
    normalized = "".join(out)
    return normalized or s


def _validate_and_normalize_customer_row(
    row: Dict[str, Any],
    *,
    row_number: int,
) -> Tuple[Optional[Dict[str, Any]], List[str]]:
    errors: List[str] = []

    customer_id = _to_int(row.get("customer_id"))
    name = str(row.get("name") or "").strip()
    if not name:
        errors.append("name is required")

    phone = _normalize_phone(row.get("phone"))
    if not str(phone or "").strip():
        errors.append("phone is required")

    email = str(row.get("email") or "").strip() or None
    vehicle_number = str(row.get("vehicle_number") or "").strip() or None

    if errors:
        return None, errors

    return (
        {
            "customer_id": customer_id,
            "name": name,
            "phone": phone,
            "email": email,
            "vehicle_number": vehicle_number,
            "_row": row_number,
        },
        [],
    )


@router.get("/customers/template")
def download_customers_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user

    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(CUSTOMERS_TEMPLATE_COLUMNS)
        writer.writerow(["", "John Doe", "+919999999999", "john@example.com", "TN-09-AA-1234"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=customers_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(CUSTOMERS_TEMPLATE_COLUMNS)
    ws.append(["", "John Doe", "+919999999999", "john@example.com", "TN-09-AA-1234"])
    notes = wb.create_sheet("Notes")
    notes.append(["How to use"])
    notes.append(["- Fill the Customers sheet, then upload in Customers → Bulk Upload."])
    notes.append(["- To UPDATE existing rows, set customer_id (or keep phone same)."])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customers_bulk_template.xlsx"},
    )


@router.post("/customers/preview")
async def preview_customers_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user

    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _cols, raw_rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _cols, raw_rows = _read_xlsx_bytes(content, sheet_name_hint="Customers")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    max_rows = 5000
    if len(raw_rows) > max_rows:
        raise HTTPException(status_code=400, detail=f"Too many rows ({len(raw_rows)}). Please upload {max_rows} rows or fewer per file.")

    errors: List[schemas.BulkRowError] = []
    normalized_rows: List[Dict[str, Any]] = []

    seen_ids: set[int] = set()
    seen_phones: set[str] = set()
    seen_emails: set[str] = set()

    for idx, row in enumerate(raw_rows, start=2):
        normalized, row_errors = _validate_and_normalize_customer_row(row, row_number=idx)
        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue

        assert normalized is not None

        cid = normalized.get("customer_id")
        if cid is not None:
            if int(cid) in seen_ids:
                errors.append(schemas.BulkRowError(row=idx, message="Duplicate customer_id in upload"))
                continue
            seen_ids.add(int(cid))

        phone = str(normalized.get("phone") or "").strip()
        if phone:
            if phone in seen_phones:
                errors.append(schemas.BulkRowError(row=idx, message="Duplicate phone in upload"))
                continue
            seen_phones.add(phone)

        email = str(normalized.get("email") or "").strip().lower()
        if email:
            if email in seen_emails:
                errors.append(schemas.BulkRowError(row=idx, message="Duplicate email in upload"))
                continue
            seen_emails.add(email)

        normalized_rows.append(normalized)

    return {
        "columns": CUSTOMERS_TEMPLATE_COLUMNS,
        "total_rows": len(raw_rows),
        "valid_rows": len(normalized_rows),
        "errors": [e.model_dump() for e in errors],
        "rows": normalized_rows,
    }


@router.post("/customers/commit")
def commit_customers_upload(
    payload: schemas.SalesBulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user

    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        normalized, errors = _validate_and_normalize_customer_row(row, row_number=int(row.get("_row") or i))
        if errors or normalized is None:
            for msg in errors or ["Invalid row"]:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message=msg))
            continue

        mode = payload.mode
        cid = normalized.get("customer_id")
        phone = str(normalized.get("phone") or "").strip()
        email = str(normalized.get("email") or "").strip() or None

        db_customer = None
        if cid is not None:
            db_customer = db.query(models.Customer).filter(models.Customer.id == int(cid)).first()
        if db_customer is None and phone:
            db_customer = db.query(models.Customer).filter(models.Customer.phone == phone).first()

        if db_customer is None and mode == schemas.SalesBulkCommitMode.UPDATE_ONLY:
            row_errors.append(
                schemas.BulkRowError(
                    row=int(normalized.get("_row") or i),
                    message="No matching customer found for update (use customer_id or existing phone)",
                )
            )
            continue

        if db_customer is None:
            # Insert
            if phone and db.query(models.Customer).filter(models.Customer.phone == phone).first():
                row_errors.append(schemas.BulkRowError(row=int(normalized.get("_row") or i), message="Phone number already registered"))
                continue
            if email and db.query(models.Customer).filter(models.Customer.email == email).first():
                row_errors.append(schemas.BulkRowError(row=int(normalized.get("_row") or i), message="Email already registered"))
                continue

            db_customer = models.Customer(
                name=str(normalized.get("name") or "").strip(),
                phone=phone,
                email=email,
                vehicle_number=normalized.get("vehicle_number"),
            )
            db.add(db_customer)
            inserted += 1
            continue

        # Update
        if mode == schemas.SalesBulkCommitMode.INSERT_ONLY:
            continue

        # Check uniqueness when changing phone/email
        if phone and phone != db_customer.phone:
            if db.query(models.Customer).filter(models.Customer.phone == phone, models.Customer.id != db_customer.id).first():
                row_errors.append(schemas.BulkRowError(row=int(normalized.get("_row") or i), message="Phone number already registered"))
                continue
            db_customer.phone = phone

        if email and email != db_customer.email:
            if db.query(models.Customer).filter(models.Customer.email == email, models.Customer.id != db_customer.id).first():
                row_errors.append(schemas.BulkRowError(row=int(normalized.get("_row") or i), message="Email already registered"))
                continue
            db_customer.email = email
        if email is None:
            db_customer.email = None

        db_customer.name = str(normalized.get("name") or "").strip()
        db_customer.vehicle_number = normalized.get("vehicle_number")
        updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Validation failed",
                "errors": [e.model_dump() for e in row_errors[:200]],
            },
        )

    db.commit()
    return {"inserted": inserted, "updated": updated, "failed": len(row_errors)}


@router.get("/products/template")
def download_products_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(PRODUCTS_TEMPLATE_COLUMNS)
        writer.writerow(["Petrol", "petrol", "true"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=products_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(PRODUCTS_TEMPLATE_COLUMNS)
    ws.append(["Petrol", "petrol", "true"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_bulk_template.xlsx"},
    )


@router.post("/products/preview", response_model=schemas.BulkPreviewResponse)
async def preview_products_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _, rows = _read_xlsx_bytes(content, sheet_name_hint="Products")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []
    category_names = _get_category_names(db)
    for idx, row in enumerate(rows, start=2):
        row_errors = _validate_required(row, ["product_name", "fuel_type"])
        fuel = _normalize_fuel_type(row.get("fuel_type"))
        if fuel and fuel not in category_names:
            row_errors.append("fuel_type must match a configured product category")
        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue

        normalized.append(
            {
                "product_name": str(row.get("product_name") or "").strip(),
                "fuel_type": fuel,
                "is_active": _to_bool(row.get("is_active")) if row.get("is_active") is not None else True,
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=PRODUCTS_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/products/commit", response_model=schemas.BulkCommitResponse)
def commit_products_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []
    category_names = _get_category_names(db)

    for i, row in enumerate(payload.rows or [], start=1):
        name = str(row.get("product_name") or "").strip()
        fuel = _normalize_fuel_type(row.get("fuel_type"))
        if not name or fuel not in category_names:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Invalid product_name/fuel_type"))
            continue

        db_product = db.query(models.Product).filter(models.Product.product_name == name).first()
        if db_product is None:
            if payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Product not found for update"))
                continue
            db_product = models.Product(product_name=name, fuel_type=fuel)
            db.add(db_product)
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            db_product.fuel_type = fuel
            updated += 1

        is_active = row.get("is_active")
        if is_active is not None:
            db_product.is_active = bool(is_active)

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return {"inserted": inserted, "updated": updated, "failed": len(row_errors)}


@router.get("/product-prices/template")
def download_product_prices_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(PRODUCT_PRICES_TEMPLATE_COLUMNS)
        writer.writerow(["", "Petrol", date.today().isoformat(), "104.5", "Opening price"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=product_prices_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "ProductPrices"
    ws.append(PRODUCT_PRICES_TEMPLATE_COLUMNS)
    ws.append(["", "Petrol", date.today().isoformat(), 104.5, "Opening price"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_prices_bulk_template.xlsx"},
    )


@router.post("/product-prices/preview", response_model=schemas.BulkPreviewResponse)
async def preview_product_prices_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _, rows = _read_xlsx_bytes(content, sheet_name_hint="ProductPrices")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        row_errors = _validate_required(row, ["effective_date", "price_per_liter"])
        product_id = _to_int(row.get("product_id"))
        product_name = str(row.get("product_name") or "").strip()
        if not product_id and not product_name:
            row_errors.append("product_id or product_name is required")
        eff = _to_date(row.get("effective_date"))
        if eff is None:
            row_errors.append("effective_date must be YYYY-MM-DD")
        price = _to_float(row.get("price_per_liter"))
        if price is None or price < 0:
            row_errors.append("price_per_liter must be >= 0")
        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue
        normalized.append(
            {
                "product_id": product_id,
                "product_name": product_name,
                "effective_date": eff.isoformat(),
                "price_per_liter": float(price),
                "remarks": (str(row.get("remarks") or "").strip() or None),
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=PRODUCT_PRICES_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/product-prices/commit", response_model=schemas.BulkCommitResponse)
def commit_product_prices_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        product_id = _to_int(row.get("product_id"))
        if product_id is None:
            name = str(row.get("product_name") or "").strip()
            product = db.query(models.Product).filter(models.Product.product_name == name).first()
            if not product:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Product not found"))
                continue
            product_id = product.id

        eff = _to_date(row.get("effective_date"))
        price = _to_float(row.get("price_per_liter"))
        if eff is None or price is None:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Invalid effective_date/price_per_liter"))
            continue

        existing = (
            db.query(models.ProductPrice)
            .filter(models.ProductPrice.product_id == product_id, models.ProductPrice.effective_date == eff)
            .first()
        )
        if existing is None:
            if payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Price row not found for update"))
                continue
            db.add(
                models.ProductPrice(
                    product_id=product_id,
                    effective_date=eff,
                    price_per_liter=float(price),
                    remarks=row.get("remarks"),
                    created_by_user_id=current_user.id,
                )
            )
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            existing.price_per_liter = float(price)
            existing.remarks = row.get("remarks")
            updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return {"inserted": inserted, "updated": updated, "failed": len(row_errors)}


@router.get("/dispensers/template")
def download_dispensers_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(DISPENSERS_TEMPLATE_COLUMNS)
        writer.writerow(["D-01", "true"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=dispensers_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Dispensers"
    ws.append(DISPENSERS_TEMPLATE_COLUMNS)
    ws.append(["D-01", "true"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dispensers_bulk_template.xlsx"},
    )


@router.post("/dispensers/preview", response_model=schemas.BulkPreviewResponse)
async def preview_dispensers_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _, rows = _read_xlsx_bytes(content, sheet_name_hint="Dispensers")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        row_errors = _validate_required(row, ["dispenser_number"])
        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue
        normalized.append(
            {
                "dispenser_number": str(row.get("dispenser_number") or "").strip(),
                "is_active": _to_bool(row.get("is_active")) if row.get("is_active") is not None else True,
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=DISPENSERS_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/dispensers/commit", response_model=schemas.BulkCommitResponse)
def commit_dispensers_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        number = str(row.get("dispenser_number") or "").strip()
        if not number:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="dispenser_number is required"))
            continue

        dispenser = db.query(models.Dispenser).filter(models.Dispenser.dispenser_number == number).first()
        if dispenser is None:
            if payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Dispenser not found for update"))
                continue
            dispenser = models.Dispenser(dispenser_number=number, fuel_type="petrol")
            db.add(dispenser)
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            updated += 1

        is_active = row.get("is_active")
        if is_active is not None:
            dispenser.is_active = bool(is_active)

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return {"inserted": inserted, "updated": updated, "failed": len(row_errors)}


@router.get("/tanks/template")
def download_tanks_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(TANKS_TEMPLATE_COLUMNS)
        writer.writerow(["T-01", "", "Petrol", "20000", "0", "false", "", "", "Main tank"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=tanks_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Tanks"
    ws.append(TANKS_TEMPLATE_COLUMNS)
    ws.append(["T-01", "", "Petrol", 20000, 0, "false", "", "", "Main tank"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tanks_bulk_template.xlsx"},
    )


@router.post("/tanks/preview", response_model=schemas.BulkPreviewResponse)
async def preview_tanks_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _, rows = _read_xlsx_bytes(content, sheet_name_hint="Tanks")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        row_errors = _validate_required(row, ["tank_name", "capacity"])
        product_id = _to_int(row.get("product_id"))
        product_name = str(row.get("product_name") or "").strip()
        if not product_id and not product_name:
            row_errors.append("product_id or product_name is required")
        capacity = _to_float(row.get("capacity"))
        if capacity is None or capacity <= 0:
            row_errors.append("capacity must be > 0")
        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue

        normalized.append(
            {
                "tank_name": str(row.get("tank_name") or "").strip(),
                "product_id": product_id,
                "product_name": product_name,
                "capacity": float(capacity),
                "current_volume": _to_float(row.get("current_volume")) or 0.0,
                "is_buffer": _to_bool(row.get("is_buffer")) if row.get("is_buffer") is not None else False,
                "calibration_date": (_to_date(row.get("calibration_date")).isoformat() if _to_date(row.get("calibration_date")) else None),
                "calibration_due_date": (_to_date(row.get("calibration_due_date")).isoformat() if _to_date(row.get("calibration_due_date")) else None),
                "remarks": (str(row.get("remarks") or "").strip() or None),
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=TANKS_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/tanks/commit", response_model=schemas.BulkCommitResponse)
def commit_tanks_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        name = str(row.get("tank_name") or "").strip()
        if not name:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="tank_name is required"))
            continue

        product_id = _to_int(row.get("product_id"))
        if product_id is None:
            product_name = str(row.get("product_name") or "").strip()
            product = db.query(models.Product).filter(models.Product.product_name == product_name).first()
            if not product:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Product not found"))
                continue
            product_id = product.id

        capacity = _to_float(row.get("capacity"))
        if capacity is None or capacity <= 0:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Invalid capacity"))
            continue

        tank = db.query(models.Tank).filter(models.Tank.tank_name == name).first()
        if tank is None:
            if payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Tank not found for update"))
                continue
            tank = models.Tank(
                tank_name=name,
                product_id=product_id,
                capacity=float(capacity),
                current_volume=float(row.get("current_volume") or 0.0),
                is_buffer=bool(row.get("is_buffer") or False),
                calibration_date=_to_date(row.get("calibration_date")),
                calibration_due_date=_to_date(row.get("calibration_due_date")),
                remarks=row.get("remarks"),
            )
            db.add(tank)
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            tank.product_id = product_id
            tank.capacity = float(capacity)
            tank.current_volume = float(row.get("current_volume") or 0.0)
            tank.is_buffer = bool(row.get("is_buffer") or False)
            tank.calibration_date = _to_date(row.get("calibration_date"))
            tank.calibration_due_date = _to_date(row.get("calibration_due_date"))
            tank.remarks = row.get("remarks")
            updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return {"inserted": inserted, "updated": updated, "failed": len(row_errors)}


@router.get("/nozzles/template")
def download_nozzles_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(NOZZLES_TEMPLATE_COLUMNS)
        writer.writerow(["", "D-01", "N-01", "", "Petrol", "", "T-01", "true"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=nozzles_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Nozzles"
    ws.append(NOZZLES_TEMPLATE_COLUMNS)
    ws.append(["", "D-01", "N-01", "", "Petrol", "", "T-01", "true"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=nozzles_bulk_template.xlsx"},
    )


@router.post("/nozzles/preview", response_model=schemas.BulkPreviewResponse)
async def preview_nozzles_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _, rows = _read_xlsx_bytes(content, sheet_name_hint="Nozzles")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        row_errors = _validate_required(row, ["nozzle_number"])
        if not (_to_int(row.get("dispenser_id")) or str(row.get("dispenser_number") or "").strip()):
            row_errors.append("dispenser_id or dispenser_number is required")
        if not (_to_int(row.get("product_id")) or str(row.get("product_name") or "").strip()):
            row_errors.append("product_id or product_name is required")
        if not (_to_int(row.get("tank_id")) or str(row.get("tank_name") or "").strip()):
            row_errors.append("tank_id or tank_name is required")
        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue

        normalized.append(
            {
                "dispenser_id": _to_int(row.get("dispenser_id")),
                "dispenser_number": str(row.get("dispenser_number") or "").strip(),
                "nozzle_number": str(row.get("nozzle_number") or "").strip(),
                "product_id": _to_int(row.get("product_id")),
                "product_name": str(row.get("product_name") or "").strip(),
                "tank_id": _to_int(row.get("tank_id")),
                "tank_name": str(row.get("tank_name") or "").strip(),
                "is_active": _to_bool(row.get("is_active")) if row.get("is_active") is not None else True,
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=NOZZLES_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/nozzles/commit", response_model=schemas.BulkCommitResponse)
def commit_nozzles_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        nozzle_number = str(row.get("nozzle_number") or "").strip()
        if not nozzle_number:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="nozzle_number is required"))
            continue

        dispenser_id = _to_int(row.get("dispenser_id"))
        if dispenser_id is None:
            disp_number = str(row.get("dispenser_number") or "").strip()
            dispenser = db.query(models.Dispenser).filter(models.Dispenser.dispenser_number == disp_number).first()
            if not dispenser:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Dispenser not found"))
                continue
            dispenser_id = dispenser.id

        product_id = _to_int(row.get("product_id"))
        if product_id is None:
            product_name = str(row.get("product_name") or "").strip()
            product = db.query(models.Product).filter(models.Product.product_name == product_name).first()
            if not product:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Product not found"))
                continue
            product_id = product.id
        product = db.query(models.Product).filter(models.Product.id == product_id).first()
        if not product:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Product not found"))
            continue

        tank_id = _to_int(row.get("tank_id"))
        if tank_id is None:
            tank_name = str(row.get("tank_name") or "").strip()
            tank = db.query(models.Tank).filter(models.Tank.tank_name == tank_name).first()
            if not tank:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Tank not found"))
                continue
            tank_id = tank.id
        tank = db.query(models.Tank).filter(models.Tank.id == tank_id).first()
        if not tank:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Tank not found"))
            continue
        if tank.product_id != product_id:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Tank product does not match nozzle product"))
            continue

        nozzle = (
            db.query(models.Nozzle)
            .filter(models.Nozzle.dispenser_id == dispenser_id, models.Nozzle.nozzle_number == nozzle_number)
            .first()
        )
        if nozzle is None:
            if payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Nozzle not found for update"))
                continue
            nozzle = models.Nozzle(
                dispenser_id=dispenser_id,
                nozzle_number=nozzle_number,
                fuel_type=product.fuel_type,
                product_id=product_id,
                tank_id=tank_id,
                is_active=bool(row.get("is_active") if row.get("is_active") is not None else True),
            )
            db.add(nozzle)
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            nozzle.product_id = product_id
            nozzle.tank_id = tank_id
            nozzle.fuel_type = product.fuel_type
            nozzle.is_active = bool(row.get("is_active") if row.get("is_active") is not None else nozzle.is_active)
            updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return {"inserted": inserted, "updated": updated, "failed": len(row_errors)}


@router.get("/meters/template")
def download_meters_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(METERS_TEMPLATE_COLUMNS)
        writer.writerow(["", "D-01", "N-01", "M-01", "99999", "0", "true"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=meters_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Meters"
    ws.append(METERS_TEMPLATE_COLUMNS)
    ws.append(["", "D-01", "N-01", "M-01", 99999, 0, "true"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=meters_bulk_template.xlsx"},
    )


@router.post("/meters/preview", response_model=schemas.BulkPreviewResponse)
async def preview_meters_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _, rows = _read_xlsx_bytes(content, sheet_name_hint="Meters")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        row_errors = _validate_required(row, ["meter_name"])
        if not (_to_int(row.get("nozzle_id")) or (row.get("dispenser_number") and row.get("nozzle_number"))):
            row_errors.append("nozzle_id or (dispenser_number + nozzle_number) is required")
        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue

        normalized.append(
            {
                "nozzle_id": _to_int(row.get("nozzle_id")),
                "dispenser_number": str(row.get("dispenser_number") or "").strip(),
                "nozzle_number": str(row.get("nozzle_number") or "").strip(),
                "meter_name": str(row.get("meter_name") or "").strip(),
                "max_value": _to_float(row.get("max_value")),
                "last_reading": _to_float(row.get("last_reading")) or 0.0,
                "is_active": _to_bool(row.get("is_active")) if row.get("is_active") is not None else True,
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=METERS_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/meters/commit", response_model=schemas.BulkCommitResponse)
def commit_meters_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        meter_name = str(row.get("meter_name") or "").strip()
        if not meter_name:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="meter_name is required"))
            continue

        nozzle_id = _to_int(row.get("nozzle_id"))
        if nozzle_id is None:
            disp_number = str(row.get("dispenser_number") or "").strip()
            nozzle_number = str(row.get("nozzle_number") or "").strip()
            nozzle = (
                db.query(models.Nozzle)
                .join(models.Dispenser, models.Nozzle.dispenser_id == models.Dispenser.id)
                .filter(models.Dispenser.dispenser_number == disp_number, models.Nozzle.nozzle_number == nozzle_number)
                .first()
            )
            if not nozzle:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Nozzle not found"))
                continue
            nozzle_id = nozzle.id

        meter = (
            db.query(models.Meter)
            .filter(models.Meter.nozzle_id == nozzle_id, models.Meter.meter_name == meter_name)
            .first()
        )
        if meter is None:
            if payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Meter not found for update"))
                continue
            meter = models.Meter(
                nozzle_id=nozzle_id,
                meter_name=meter_name,
                max_value=_to_float(row.get("max_value")),
                last_reading=float(row.get("last_reading") or 0.0),
                is_active=bool(row.get("is_active") if row.get("is_active") is not None else True),
            )
            db.add(meter)
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            meter.max_value = _to_float(row.get("max_value"))
            meter.last_reading = float(row.get("last_reading") or 0.0)
            meter.is_active = bool(row.get("is_active") if row.get("is_active") is not None else meter.is_active)
            updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return {"inserted": inserted, "updated": updated, "failed": len(row_errors)}


@router.get("/employees/template")
def download_employees_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(EMPLOYEES_TEMPLATE_COLUMNS)
        writer.writerow(["John Doe", "1990-01-01", "Address", "9999999999", "ID-001", "", "Cashier", "true"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=employees_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append(EMPLOYEES_TEMPLATE_COLUMNS)
    ws.append(["John Doe", "1990-01-01", "Address", "9999999999", "ID-001", "", "Cashier", "true"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employees_bulk_template.xlsx"},
    )


@router.post("/employees/preview", response_model=schemas.BulkPreviewResponse)
async def preview_employees_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _, rows = _read_xlsx_bytes(content, sheet_name_hint="Employees")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        row_errors = _validate_required(row, ["employee_name"])
        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue

        normalized.append(
            {
                "employee_name": str(row.get("employee_name") or "").strip(),
                "dob": (_to_date(row.get("dob")).isoformat() if _to_date(row.get("dob")) else None),
                "address": (str(row.get("address") or "").strip() or None),
                "contact_no": (str(row.get("contact_no") or "").strip() or None),
                "id_no": (str(row.get("id_no") or "").strip() or None),
                "designation_id": _to_int(row.get("designation_id")),
                "designation_name": str(row.get("designation_name") or "").strip(),
                "is_active": _to_bool(row.get("is_active")) if row.get("is_active") is not None else True,
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=EMPLOYEES_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/employees/commit", response_model=schemas.BulkCommitResponse)
def commit_employees_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        name = str(row.get("employee_name") or "").strip()
        if not name:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="employee_name is required"))
            continue

        designation_id = _to_int(row.get("designation_id"))
        if designation_id is None and row.get("designation_name"):
            d = db.query(models.Designation).filter(models.Designation.name == row.get("designation_name")).first()
            if not d:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Designation not found"))
                continue
            designation_id = d.id

        id_no = str(row.get("id_no") or "").strip() or None
        contact_no = str(row.get("contact_no") or "").strip() or None
        employee = None
        if id_no:
            employee = db.query(models.Employee).filter(models.Employee.id_no == id_no).first()
        if employee is None and contact_no:
            employee = db.query(models.Employee).filter(models.Employee.employee_name == name, models.Employee.contact_no == contact_no).first()

        if employee is None:
            if payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Employee not found for update"))
                continue
            employee = models.Employee(
                employee_name=name,
                dob=_to_date(row.get("dob")),
                address=row.get("address"),
                contact_no=contact_no,
                id_no=id_no,
                designation_id=designation_id,
                is_active=bool(row.get("is_active") if row.get("is_active") is not None else True),
            )
            db.add(employee)
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            employee.employee_name = name
            employee.dob = _to_date(row.get("dob"))
            employee.address = row.get("address")
            employee.contact_no = contact_no
            employee.id_no = id_no
            employee.designation_id = designation_id
            employee.is_active = bool(row.get("is_active") if row.get("is_active") is not None else employee.is_active)
            updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return {"inserted": inserted, "updated": updated, "failed": len(row_errors)}


@router.get("/designations/template")
def download_designations_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(DESIGNATIONS_TEMPLATE_COLUMNS)
        writer.writerow(["Cashier", "true"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=designations_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Designations"
    ws.append(DESIGNATIONS_TEMPLATE_COLUMNS)
    ws.append(["Cashier", "true"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=designations_bulk_template.xlsx"},
    )


@router.post("/designations/preview", response_model=schemas.BulkPreviewResponse)
async def preview_designations_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _, rows = _read_xlsx_bytes(content, sheet_name_hint="Designations")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        row_errors = _validate_required(row, ["name"])
        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue
        normalized.append(
            {
                "name": str(row.get("name") or "").strip(),
                "is_active": _to_bool(row.get("is_active")) if row.get("is_active") is not None else True,
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=DESIGNATIONS_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/designations/commit", response_model=schemas.BulkCommitResponse)
def commit_designations_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        name = str(row.get("name") or "").strip()
        if not name:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="name is required"))
            continue

        designation = db.query(models.Designation).filter(models.Designation.name == name).first()
        if designation is None:
            if payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Designation not found for update"))
                continue
            designation = models.Designation(name=name, is_active=bool(row.get("is_active") if row.get("is_active") is not None else True))
            db.add(designation)
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            designation.is_active = bool(row.get("is_active") if row.get("is_active") is not None else designation.is_active)
            updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return {"inserted": inserted, "updated": updated, "failed": len(row_errors)}


@router.get("/shifts/template")
def download_shifts_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(SHIFTS_TEMPLATE_COLUMNS)
        writer.writerow(["A", "06:00:00", "14:00:00", "true", "Morning shift"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=shifts_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Shifts"
    ws.append(SHIFTS_TEMPLATE_COLUMNS)
    ws.append(["A", "06:00:00", "14:00:00", "true", "Morning shift"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=shifts_bulk_template.xlsx"},
    )


@router.post("/shifts/preview", response_model=schemas.BulkPreviewResponse)
async def preview_shifts_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _, rows = _read_xlsx_bytes(content, sheet_name_hint="Shifts")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        row_errors = _validate_required(row, ["shift", "start_time", "end_time"])
        shift = str(row.get("shift") or "").strip().upper()
        if shift not in ("A", "B", "C"):
            row_errors.append("shift must be A, B, or C")
        start_time = _to_time(row.get("start_time"))
        end_time = _to_time(row.get("end_time"))
        if start_time is None or end_time is None:
            row_errors.append("start_time/end_time must be HH:MM:SS")
        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue
        normalized.append(
            {
                "shift": shift,
                "start_time": start_time.isoformat(),
                "end_time": end_time.isoformat(),
                "is_active": _to_bool(row.get("is_active")) if row.get("is_active") is not None else True,
                "remarks": (str(row.get("remarks") or "").strip() or None),
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=SHIFTS_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/shifts/commit", response_model=schemas.BulkCommitResponse)
def commit_shifts_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        shift = str(row.get("shift") or "").strip().upper()
        start_time = _to_time(row.get("start_time"))
        end_time = _to_time(row.get("end_time"))
        if shift not in ("A", "B", "C") or start_time is None or end_time is None:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Invalid shift/start_time/end_time"))
            continue

        cfg = db.query(models.ShiftConfig).filter(models.ShiftConfig.shift == models.ShiftCode(shift)).first()
        if cfg is None:
            if payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Shift not found for update"))
                continue
            cfg = models.ShiftConfig(
                shift=models.ShiftCode(shift),
                start_time=start_time,
                end_time=end_time,
                is_active=bool(row.get("is_active") if row.get("is_active") is not None else True),
                remarks=row.get("remarks"),
            )
            db.add(cfg)
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            cfg.start_time = start_time
            cfg.end_time = end_time
            cfg.is_active = bool(row.get("is_active") if row.get("is_active") is not None else cfg.is_active)
            cfg.remarks = row.get("remarks")
            updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return {"inserted": inserted, "updated": updated, "failed": len(row_errors)}


@router.get("/tank-dips/template")
def download_tank_dips_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(TANK_DIPS_TEMPLATE_COLUMNS)
        writer.writerow(["", "T-01", date.today().isoformat(), "opening", "120", "", "2500", "false"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=tank_dips_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "TankDips"
    ws.append(TANK_DIPS_TEMPLATE_COLUMNS)
    ws.append(["", "T-01", date.today().isoformat(), "opening", 120, "", 2500, "false"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tank_dips_bulk_template.xlsx"},
    )


@router.post("/tank-dips/preview", response_model=schemas.BulkPreviewResponse)
async def preview_tank_dips_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _, rows = _read_xlsx_bytes(content, sheet_name_hint="TankDips")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        row_errors = _validate_required(row, ["business_date", "dip_type", "dips_mm"])
        if not (_to_int(row.get("tank_id")) or str(row.get("tank_name") or "").strip()):
            row_errors.append("tank_id or tank_name is required")
        dip_type = str(row.get("dip_type") or "").strip().lower()
        if dip_type not in ("opening", "closing"):
            row_errors.append("dip_type must be opening or closing")
        if _to_date(row.get("business_date")) is None:
            row_errors.append("business_date must be YYYY-MM-DD")
        if _to_float(row.get("dips_mm")) is None:
            row_errors.append("dips_mm must be a number")
        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue

        normalized.append(
            {
                "tank_id": _to_int(row.get("tank_id")),
                "tank_name": str(row.get("tank_name") or "").strip(),
                "business_date": _to_date(row.get("business_date")).isoformat(),
                "dip_type": dip_type,
                "dips_mm": _to_float(row.get("dips_mm")),
                "computed_volume_litres": _to_float(row.get("computed_volume_litres")),
                "manual_volume_litres": _to_float(row.get("manual_volume_litres")),
                "is_auto": _to_bool(row.get("is_auto")) if row.get("is_auto") is not None else False,
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=TANK_DIPS_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/tank-dips/commit", response_model=schemas.BulkCommitResponse)
def commit_tank_dips_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        tank_id = _to_int(row.get("tank_id"))
        if tank_id is None:
            tank_name = str(row.get("tank_name") or "").strip()
            tank = db.query(models.Tank).filter(models.Tank.tank_name == tank_name).first()
            if not tank:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Tank not found"))
                continue
            tank_id = tank.id

        business_date = _to_date(row.get("business_date"))
        dip_type_raw = str(row.get("dip_type") or "").strip().lower()
        if business_date is None or dip_type_raw not in ("opening", "closing"):
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Invalid business_date/dip_type"))
            continue
        dips_mm = _to_float(row.get("dips_mm"))
        if dips_mm is None:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Invalid dips_mm"))
            continue

        dip_type = models.TankDipType.OPENING if dip_type_raw == "opening" else models.TankDipType.CLOSING
        existing = (
            db.query(models.TankDipReading)
            .filter(
                models.TankDipReading.tank_id == tank_id,
                models.TankDipReading.business_date == business_date,
                models.TankDipReading.dip_type == dip_type,
            )
            .first()
        )

        computed_vol = _to_float(row.get("computed_volume_litres"))
        if computed_vol is None:
            points = (
                db.query(models.TankCalibrationPoint)
                .filter(models.TankCalibrationPoint.tank_id == tank_id)
                .order_by(models.TankCalibrationPoint.dips_mm.asc())
                .all()
            )
            try:
                computed_vol = _interpolate_volume(points, float(dips_mm))
            except Exception:
                computed_vol = None

        if existing is None:
            if payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Dip reading not found for update"))
                continue
            db.add(
                models.TankDipReading(
                    tank_id=tank_id,
                    business_date=business_date,
                    dip_type=dip_type,
                    dips_mm=float(dips_mm),
                    computed_volume_litres=computed_vol,
                    manual_volume_litres=_to_float(row.get("manual_volume_litres")),
                    is_auto=bool(row.get("is_auto") or False),
                    created_by_user_id=current_user.id,
                )
            )
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            existing.dips_mm = float(dips_mm)
            existing.computed_volume_litres = computed_vol
            existing.manual_volume_litres = _to_float(row.get("manual_volume_litres"))
            existing.is_auto = bool(row.get("is_auto") or False)
            updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return {"inserted": inserted, "updated": updated, "failed": len(row_errors)}


# -----------------------------
# Tank Calibration Points Bulk Upload
# -----------------------------

@router.get("/tank-calibration-points/template")
def download_tank_calibration_points_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(TANK_CALIBRATION_POINTS_TEMPLATE_COLUMNS)
        writer.writerow(["", 1, 0, 0.0])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=tank_calibration_points_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Tank Calibration Points"
    ws.append(TANK_CALIBRATION_POINTS_TEMPLATE_COLUMNS)
    ws.append(["", 1, 0, 0.0])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tank_calibration_points_bulk_template.xlsx"},
    )


@router.post("/tank-calibration-points/preview", response_model=schemas.BulkPreviewResponse)
async def preview_tank_calibration_points_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _cols, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _cols, rows = _read_xlsx_bytes(content, sheet_name_hint="Tank Calibration Points")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []
    seen_pairs: set[Tuple[int, float]] = set()

    for idx, row in enumerate(rows, start=2):
        point_id = _to_int(row.get("calibration_point_id"))
        tank_id = _to_int(row.get("tank_id"))
        dips_mm = _to_float(row.get("dips_mm"))
        volume = _to_float(row.get("volume_in_litres"))
        row_errors: List[str] = []
        if tank_id is None:
            row_errors.append("tank_id is required")
        if dips_mm is None:
            row_errors.append("dips_mm is required")
        if volume is None:
            row_errors.append("volume_in_litres is required")

        if tank_id is not None and dips_mm is not None:
            pair = (int(tank_id), float(dips_mm))
            if pair in seen_pairs:
                row_errors.append("Duplicate tank_id + dips_mm in upload")
            else:
                seen_pairs.add(pair)

        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue

        normalized.append(
            {
                "calibration_point_id": point_id,
                "tank_id": int(tank_id),
                "dips_mm": float(dips_mm),
                "volume_in_litres": float(volume),
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=TANK_CALIBRATION_POINTS_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/tank-calibration-points/commit", response_model=schemas.BulkCommitResponse)
def commit_tank_calibration_points_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        point_id = _to_int(row.get("calibration_point_id"))
        tank_id = _to_int(row.get("tank_id"))
        dips_mm = _to_float(row.get("dips_mm"))
        volume = _to_float(row.get("volume_in_litres"))
        if tank_id is None or dips_mm is None or volume is None:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="tank_id, dips_mm, volume_in_litres are required"))
            continue

        tank = db.query(models.Tank).filter(models.Tank.id == int(tank_id)).first()
        if not tank:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Tank not found"))
            continue

        db_point = None
        if point_id is not None:
            db_point = db.query(models.TankCalibrationPoint).filter(models.TankCalibrationPoint.id == int(point_id)).first()
        if db_point is None:
            db_point = (
                db.query(models.TankCalibrationPoint)
                .filter(models.TankCalibrationPoint.tank_id == int(tank_id), models.TankCalibrationPoint.dips_mm == float(dips_mm))
                .first()
            )

        if db_point is None and payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Calibration point not found for update"))
            continue

        if db_point is None:
            db.add(
                models.TankCalibrationPoint(
                    tank_id=int(tank_id),
                    dips_mm=float(dips_mm),
                    volume_in_litres=float(volume),
                )
            )
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            db_point.tank_id = int(tank_id)
            db_point.dips_mm = float(dips_mm)
            db_point.volume_in_litres = float(volume)
            updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return schemas.BulkCommitResponse(inserted=inserted, updated=updated, failed=len(row_errors))


# -----------------------------
# Fuel Inventory Bulk Upload
# -----------------------------

@router.get("/fuel-inventory/template")
def download_fuel_inventory_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(FUEL_INVENTORY_TEMPLATE_COLUMNS)
        writer.writerow(["petrol", "set", 0, 100, 0, "Initial stock set"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=fuel_inventory_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Fuel Inventory"
    ws.append(FUEL_INVENTORY_TEMPLATE_COLUMNS)
    ws.append(["petrol", "set", 0, 100, 0, "Initial stock set"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fuel_inventory_bulk_template.xlsx"},
    )


@router.post("/fuel-inventory/preview", response_model=schemas.BulkPreviewResponse)
async def preview_fuel_inventory_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _cols, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _cols, rows = _read_xlsx_bytes(content, sheet_name_hint="Fuel Inventory")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []
    category_names = _get_category_names(db)

    for idx, row in enumerate(rows, start=2):
        fuel_raw = _normalize_fuel_type(row.get("fuel_type"))
        action_raw = str(row.get("action") or "").strip().lower()
        qty = _to_float(row.get("quantity"))
        price_per_liter = _to_float(row.get("price_per_liter"))
        reorder_level = _to_float(row.get("reorder_level"))
        notes = str(row.get("notes") or "").strip() or None
        row_errors: List[str] = []

        if fuel_raw not in category_names:
            row_errors.append("fuel_type must match a configured product category")
        if action_raw not in ("set", "add", "subtract"):
            row_errors.append("action must be set, add, or subtract")
        if qty is None:
            row_errors.append("quantity is required")
        if price_per_liter is not None and price_per_liter < 0:
            row_errors.append("price_per_liter must be >= 0")
        if reorder_level is not None and reorder_level < 0:
            row_errors.append("reorder_level must be >= 0")

        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue

        db_inv = db.query(models.FuelInventory).filter(models.FuelInventory.fuel_type == fuel_raw).first()
        current_stock = float(db_inv.current_stock or 0.0) if db_inv else 0.0
        delta = float(qty or 0.0)
        if action_raw == "subtract":
            delta = -abs(delta)
        elif action_raw == "add":
            delta = abs(delta)
        else:
            delta = float(qty or 0.0) - current_stock
        new_stock = float(current_stock + delta)
        if new_stock < 0:
            errors.append(schemas.BulkRowError(row=idx, message="new_stock would become negative"))
            continue

        normalized.append(
            {
                "fuel_type": fuel_raw,
                "action": action_raw,
                "quantity": float(qty or 0.0),
                "price_per_liter": float(price_per_liter) if price_per_liter is not None else None,
                "reorder_level": float(reorder_level) if reorder_level is not None else None,
                "notes": notes,
                "current_stock": float(current_stock),
                "new_stock": float(new_stock),
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=FUEL_INVENTORY_PREVIEW_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/fuel-inventory/commit", response_model=schemas.BulkCommitResponse)
def commit_fuel_inventory_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []
    category_names = _get_category_names(db)

    for i, row in enumerate(payload.rows or [], start=1):
        fuel_raw = _normalize_fuel_type(row.get("fuel_type"))
        action_raw = str(row.get("action") or "").strip().lower()
        qty = _to_float(row.get("quantity"))
        price_per_liter = _to_float(row.get("price_per_liter"))
        reorder_level = _to_float(row.get("reorder_level"))
        notes = str(row.get("notes") or "").strip() or None
        if fuel_raw not in category_names:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Invalid fuel_type"))
            continue
        if action_raw not in ("set", "add", "subtract") or qty is None:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Invalid action/quantity"))
            continue
        if price_per_liter is not None and price_per_liter < 0:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="price_per_liter must be >= 0"))
            continue
        if reorder_level is not None and reorder_level < 0:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="reorder_level must be >= 0"))
            continue

        db_inv = db.query(models.FuelInventory).filter(models.FuelInventory.fuel_type == fuel_raw).first()
        if db_inv is None:
            if payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Fuel inventory not found for update"))
                continue
            db_inv = models.FuelInventory(
                fuel_type=fuel_raw,
                current_stock=0.0,
                price_per_liter=float(price_per_liter or 0.0),
                reorder_level=float(reorder_level or 0.0),
            )
            db.add(db_inv)
            db.flush()
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue

        current_stock = float(db_inv.current_stock or 0.0)
        if action_raw == "set":
            new_stock = float(qty)
            delta = float(new_stock - current_stock)
        elif action_raw == "add":
            delta = abs(float(qty))
            new_stock = current_stock + delta
        else:
            delta = -abs(float(qty))
            new_stock = current_stock + delta

        if new_stock < 0:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="new_stock would become negative"))
            continue

        db_inv.current_stock = float(new_stock)
        if price_per_liter is not None:
            db_inv.price_per_liter = float(price_per_liter)
        if reorder_level is not None:
            db_inv.reorder_level = float(reorder_level)

        db.add(
            models.InventoryLog(
                fuel_type=fuel_raw,
                action="inventory_bulk_adjust",
                quantity=float(delta),
                previous_stock=current_stock,
                new_stock=float(new_stock),
                notes=notes or f"Bulk inventory {action_raw}",
            )
        )
        updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return schemas.BulkCommitResponse(inserted=inserted, updated=updated, failed=len(row_errors))


# -----------------------------
# Tank Transfers Bulk Upload
# -----------------------------

@router.get("/tank-transfers/template")
def download_tank_transfers_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(TANK_TRANSFERS_TEMPLATE_COLUMNS)
        writer.writerow(["", 1, 2, 1, 100, "manual", ""])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=tank_transfers_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Tank Transfers"
    ws.append(TANK_TRANSFERS_TEMPLATE_COLUMNS)
    ws.append(["", 1, 2, 1, 100, "manual", ""])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tank_transfers_bulk_template.xlsx"},
    )


@router.post("/tank-transfers/preview", response_model=schemas.BulkPreviewResponse)
async def preview_tank_transfers_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _cols, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _cols, rows = _read_xlsx_bytes(content, sheet_name_hint="Tank Transfers")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []

    for idx, row in enumerate(rows, start=2):
        transfer_id = _to_int(row.get("transfer_id"))
        from_tank_id = _to_int(row.get("from_tank_id"))
        to_tank_id = _to_int(row.get("to_tank_id"))
        product_id = _to_int(row.get("product_id"))
        volume = _to_float(row.get("volume"))
        transfer_type_raw = str(row.get("transfer_type") or "manual").strip().lower()
        user_id = _to_int(row.get("user_id"))

        row_errors: List[str] = []
        if from_tank_id is None or to_tank_id is None or product_id is None:
            row_errors.append("from_tank_id, to_tank_id, product_id are required")
        if volume is None or volume <= 0:
            row_errors.append("volume must be > 0")
        if transfer_type_raw not in ("manual", "testing_to_buffer", "buffer_to_main"):
            row_errors.append("transfer_type must be manual, testing_to_buffer, or buffer_to_main")
        if from_tank_id is not None and to_tank_id is not None and int(from_tank_id) == int(to_tank_id):
            row_errors.append("from_tank_id and to_tank_id must be different")

        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue

        normalized.append(
            {
                "transfer_id": transfer_id,
                "from_tank_id": int(from_tank_id),
                "to_tank_id": int(to_tank_id),
                "product_id": int(product_id),
                "volume": float(volume),
                "transfer_type": transfer_type_raw,
                "user_id": user_id,
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=TANK_TRANSFERS_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/tank-transfers/commit", response_model=schemas.BulkCommitResponse)
def commit_tank_transfers_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        transfer_id = _to_int(row.get("transfer_id"))
        from_tank_id = _to_int(row.get("from_tank_id"))
        to_tank_id = _to_int(row.get("to_tank_id"))
        product_id = _to_int(row.get("product_id"))
        volume = _to_float(row.get("volume"))
        transfer_type_raw = str(row.get("transfer_type") or "manual").strip().lower()
        user_id = _to_int(row.get("user_id")) or current_user.id

        if (
            from_tank_id is None
            or to_tank_id is None
            or product_id is None
            or volume is None
            or volume <= 0
            or transfer_type_raw not in ("manual", "testing_to_buffer", "buffer_to_main")
        ):
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Invalid transfer data"))
            continue

        from_tank = db.query(models.Tank).filter(models.Tank.id == int(from_tank_id)).first()
        to_tank = db.query(models.Tank).filter(models.Tank.id == int(to_tank_id)).first()
        product = db.query(models.Product).filter(models.Product.id == int(product_id)).first()
        user = db.query(models.User).filter(models.User.id == int(user_id)).first()
        if not from_tank or not to_tank or not product or not user:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Tank/product/user not found"))
            continue

        db_transfer = None
        if transfer_id is not None:
            db_transfer = db.query(models.TankTransfer).filter(models.TankTransfer.id == int(transfer_id)).first()

        if db_transfer is None and payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Transfer not found for update"))
            continue

        if db_transfer is None:
            db.add(
                models.TankTransfer(
                    from_tank_id=int(from_tank_id),
                    to_tank_id=int(to_tank_id),
                    product_id=int(product_id),
                    volume=float(volume),
                    transfer_type=models.TankTransferType(transfer_type_raw),
                    user_id=int(user_id),
                )
            )
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            db_transfer.from_tank_id = int(from_tank_id)
            db_transfer.to_tank_id = int(to_tank_id)
            db_transfer.product_id = int(product_id)
            db_transfer.volume = float(volume)
            db_transfer.transfer_type = models.TankTransferType(transfer_type_raw)
            db_transfer.user_id = int(user_id)
            updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return schemas.BulkCommitResponse(inserted=inserted, updated=updated, failed=len(row_errors))


# -----------------------------
# Dispenser Shift Assignments Bulk Upload
# -----------------------------

@router.get("/dispenser-shift-assignments/template")
def download_dispenser_shift_assignments_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(DISPENSER_SHIFT_ASSIGNMENTS_TEMPLATE_COLUMNS)
        writer.writerow(["", date.today().isoformat(), "A", 1, 1])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=dispenser_shift_assignments_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Dispenser Shift Assignments"
    ws.append(DISPENSER_SHIFT_ASSIGNMENTS_TEMPLATE_COLUMNS)
    ws.append(["", date.today().isoformat(), "A", 1, 1])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dispenser_shift_assignments_bulk_template.xlsx"},
    )


@router.post("/dispenser-shift-assignments/preview", response_model=schemas.BulkPreviewResponse)
async def preview_dispenser_shift_assignments_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _cols, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _cols, rows = _read_xlsx_bytes(content, sheet_name_hint="Dispenser Shift Assignments")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []

    for idx, row in enumerate(rows, start=2):
        assignment_id = _to_int(row.get("assignment_id"))
        business_date = _to_date(row.get("business_date"))
        shift_raw = str(row.get("shift") or "").strip().upper()
        dispenser_id = _to_int(row.get("dispenser_id"))
        operator_id = _to_int(row.get("operator_id"))
        row_errors: List[str] = []

        if business_date is None:
            row_errors.append("business_date is required (YYYY-MM-DD)")
        if shift_raw not in ("A", "B", "C"):
            row_errors.append("shift must be A, B, or C")
        if dispenser_id is None or operator_id is None:
            row_errors.append("dispenser_id and operator_id are required")

        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue

        normalized.append(
            {
                "assignment_id": assignment_id,
                "business_date": business_date.isoformat(),
                "shift": shift_raw,
                "dispenser_id": int(dispenser_id),
                "operator_id": int(operator_id),
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=DISPENSER_SHIFT_ASSIGNMENTS_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/dispenser-shift-assignments/commit", response_model=schemas.BulkCommitResponse)
def commit_dispenser_shift_assignments_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        assignment_id = _to_int(row.get("assignment_id"))
        business_date = _to_date(row.get("business_date"))
        shift_raw = str(row.get("shift") or "").strip().upper()
        dispenser_id = _to_int(row.get("dispenser_id"))
        operator_id = _to_int(row.get("operator_id"))
        if business_date is None or shift_raw not in ("A", "B", "C") or dispenser_id is None or operator_id is None:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Invalid assignment data"))
            continue

        dispenser = db.query(models.Dispenser).filter(models.Dispenser.id == int(dispenser_id)).first()
        operator = db.query(models.User).filter(models.User.id == int(operator_id)).first()
        if not dispenser or not operator:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Dispenser/operator not found"))
            continue

        db_assignment = None
        if assignment_id is not None:
            db_assignment = db.query(models.DispenserShiftAssignment).filter(models.DispenserShiftAssignment.id == int(assignment_id)).first()
        if db_assignment is None:
            db_assignment = (
                db.query(models.DispenserShiftAssignment)
                .filter(
                    models.DispenserShiftAssignment.business_date == business_date,
                    models.DispenserShiftAssignment.shift == models.ShiftCode(shift_raw),
                    models.DispenserShiftAssignment.dispenser_id == int(dispenser_id),
                )
                .first()
            )

        if db_assignment is None and payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Assignment not found for update"))
            continue

        if db_assignment is None:
            db.add(
                models.DispenserShiftAssignment(
                    business_date=business_date,
                    shift=models.ShiftCode(shift_raw),
                    dispenser_id=int(dispenser_id),
                    operator_id=int(operator_id),
                )
            )
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            db_assignment.business_date = business_date
            db_assignment.shift = models.ShiftCode(shift_raw)
            db_assignment.dispenser_id = int(dispenser_id)
            db_assignment.operator_id = int(operator_id)
            updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return schemas.BulkCommitResponse(inserted=inserted, updated=updated, failed=len(row_errors))


# -----------------------------
# Daily Closes Bulk Upload
# -----------------------------

@router.get("/daily-closes/template")
def download_daily_closes_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(DAILY_CLOSES_TEMPLATE_COLUMNS)
        writer.writerow(["", date.today().isoformat(), 1, 1000, 900, "Notes"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=daily_closes_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Closes"
    ws.append(DAILY_CLOSES_TEMPLATE_COLUMNS)
    ws.append(["", date.today().isoformat(), 1, 1000, 900, "Notes"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=daily_closes_bulk_template.xlsx"},
    )


@router.post("/daily-closes/preview", response_model=schemas.BulkPreviewResponse)
async def preview_daily_closes_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _cols, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _cols, rows = _read_xlsx_bytes(content, sheet_name_hint="Daily Closes")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []

    for idx, row in enumerate(rows, start=2):
        daily_close_id = _to_int(row.get("daily_close_id"))
        business_date = _to_date(row.get("business_date"))
        user_id = _to_int(row.get("user_id"))
        opening_cash = _to_float(row.get("opening_cash"))
        closing_cash = _to_float(row.get("closing_cash"))
        notes = str(row.get("notes") or "").strip() or None
        row_errors: List[str] = []

        if business_date is None:
            row_errors.append("business_date is required (YYYY-MM-DD)")
        if user_id is None:
            row_errors.append("user_id is required")

        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue

        normalized.append(
            {
                "daily_close_id": daily_close_id,
                "business_date": business_date.isoformat(),
                "user_id": int(user_id),
                "opening_cash": float(opening_cash or 0.0),
                "closing_cash": float(closing_cash or 0.0),
                "notes": notes,
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=DAILY_CLOSES_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/daily-closes/commit", response_model=schemas.BulkCommitResponse)
def commit_daily_closes_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        daily_close_id = _to_int(row.get("daily_close_id"))
        business_date = _to_date(row.get("business_date"))
        user_id = _to_int(row.get("user_id"))
        opening_cash = _to_float(row.get("opening_cash")) or 0.0
        closing_cash = _to_float(row.get("closing_cash")) or 0.0
        notes = str(row.get("notes") or "").strip() or None

        if business_date is None or user_id is None:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Invalid daily close data"))
            continue

        user = db.query(models.User).filter(models.User.id == int(user_id)).first()
        if not user:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="User not found"))
            continue

        db_close = None
        if daily_close_id is not None:
            db_close = db.query(models.DailyClose).filter(models.DailyClose.id == int(daily_close_id)).first()
        if db_close is None:
            db_close = (
                db.query(models.DailyClose)
                .filter(models.DailyClose.business_date == business_date, models.DailyClose.user_id == int(user_id))
                .first()
            )

        if db_close is None and payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Daily close not found for update"))
            continue

        if db_close is None:
            db.add(
                models.DailyClose(
                    business_date=business_date,
                    user_id=int(user_id),
                    opening_cash=float(opening_cash),
                    closing_cash=float(closing_cash),
                    notes=notes,
                )
            )
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            db_close.business_date = business_date
            db_close.user_id = int(user_id)
            db_close.opening_cash = float(opening_cash)
            db_close.closing_cash = float(closing_cash)
            db_close.notes = notes
            updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return schemas.BulkCommitResponse(inserted=inserted, updated=updated, failed=len(row_errors))


# -----------------------------
# Users Bulk Upload
# -----------------------------

@router.get("/users/template")
def download_users_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    if format == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(USERS_TEMPLATE_COLUMNS)
        writer.writerow(["", "operator1", "operator1@example.com", "ChangeMe123", "Operator One", "operator", "true"])
        content = out.getvalue().encode("utf-8")
        return StreamingResponse(
            BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=users_bulk_template.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(USERS_TEMPLATE_COLUMNS)
    ws.append(["", "operator1", "operator1@example.com", "ChangeMe123", "Operator One", "operator", "true"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_bulk_template.xlsx"},
    )


@router.post("/users/preview", response_model=schemas.BulkPreviewResponse)
async def preview_users_upload(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    if filename.endswith(".csv"):
        _cols, rows = _read_csv_bytes(content)
    elif filename.endswith(".xlsx"):
        _cols, rows = _read_xlsx_bytes(content, sheet_name_hint="Users")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    errors: List[schemas.BulkRowError] = []
    normalized: List[Dict[str, Any]] = []
    seen_usernames: set[str] = set()
    seen_emails: set[str] = set()

    for idx, row in enumerate(rows, start=2):
        user_id = _to_int(row.get("user_id"))
        username = str(row.get("username") or "").strip()
        email = str(row.get("email") or "").strip().lower()
        password = str(row.get("password") or "").strip()
        full_name = str(row.get("full_name") or "").strip() or None
        role_raw = str(row.get("role") or "").strip().lower() or None
        if role_raw == "cashier":
            role_raw = "operator"
        is_active = _to_bool(row.get("is_active")) if row.get("is_active") is not None else True

        row_errors: List[str] = []
        if not username:
            row_errors.append("username is required")
        if not email:
            row_errors.append("email is required")
        if role_raw and role_raw not in ("admin", "manager", "operator"):
            row_errors.append("role must be admin, manager, or operator")
        if password:
            if len(password.encode("utf-8")) > 72:
                row_errors.append("password is too long (max 72 bytes)")

        if username:
            if username in seen_usernames:
                row_errors.append("Duplicate username in upload")
            else:
                seen_usernames.add(username)
        if email:
            if email in seen_emails:
                row_errors.append("Duplicate email in upload")
            else:
                seen_emails.add(email)

        if row_errors:
            for msg in row_errors:
                errors.append(schemas.BulkRowError(row=idx, message=msg))
            continue

        normalized.append(
            {
                "user_id": user_id,
                "username": username,
                "email": email,
                "password": password or None,
                "full_name": full_name,
                "role": role_raw,
                "is_active": bool(is_active),
                "_row": idx,
            }
        )

    return schemas.BulkPreviewResponse(
        columns=USERS_TEMPLATE_COLUMNS,
        total_rows=len(rows),
        valid_rows=len(normalized),
        errors=errors,
        rows=normalized,
    )


@router.post("/users/commit", response_model=schemas.BulkCommitResponse)
def commit_users_upload(
    payload: schemas.BulkCommitRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    inserted = 0
    updated = 0
    row_errors: List[schemas.BulkRowError] = []

    for i, row in enumerate(payload.rows or [], start=1):
        user_id = _to_int(row.get("user_id"))
        username = str(row.get("username") or "").strip()
        email = str(row.get("email") or "").strip().lower()
        password = str(row.get("password") or "").strip()
        full_name = str(row.get("full_name") or "").strip() or None
        role_raw = str(row.get("role") or "").strip().lower() or None
        if role_raw == "cashier":
            role_raw = "operator"
        is_active = _to_bool(row.get("is_active")) if row.get("is_active") is not None else True

        if not username or not email or (role_raw and role_raw not in ("admin", "manager", "operator")):
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Invalid user data"))
            continue
        if password and len(password.encode("utf-8")) > 72:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="password is too long (max 72 bytes)"))
            continue

        db_user = None
        if user_id is not None:
            db_user = db.query(models.User).filter(models.User.id == int(user_id)).first()
        if db_user is None:
            db_user = db.query(models.User).filter(models.User.username == username).first()
        if db_user is None:
            db_user = db.query(models.User).filter(models.User.email == email).first()

        if db_user is None and payload.mode == schemas.BulkCommitMode.UPDATE_ONLY:
            row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="User not found for update"))
            continue

        if db_user is None:
            if not password:
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="password is required for new users"))
                continue
            if db.query(models.User).filter(models.User.username == username).first():
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Username already registered"))
                continue
            if db.query(models.User).filter(models.User.email == email).first():
                row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Email already registered"))
                continue

            db.add(
                models.User(
                    username=username,
                    email=email,
                    hashed_password=get_password_hash(password),
                    full_name=full_name,
                    role=models.UserRole(role_raw or "operator"),
                    is_active=bool(is_active),
                )
            )
            inserted += 1
        else:
            if payload.mode == schemas.BulkCommitMode.INSERT_ONLY:
                continue
            if username != db_user.username:
                if db.query(models.User).filter(models.User.username == username, models.User.id != db_user.id).first():
                    row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Username already registered"))
                    continue
            if email != db_user.email:
                if db.query(models.User).filter(models.User.email == email, models.User.id != db_user.id).first():
                    row_errors.append(schemas.BulkRowError(row=int(row.get("_row") or i), message="Email already registered"))
                    continue
            db_user.username = username
            db_user.email = email
            db_user.full_name = full_name
            if role_raw:
                db_user.role = models.UserRole(role_raw)
            db_user.is_active = bool(is_active)
            if password:
                db_user.hashed_password = get_password_hash(password)
            updated += 1

    if row_errors and not payload.allow_partial:
        db.rollback()
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": [e.model_dump() for e in row_errors[:200]]})

    db.commit()
    return schemas.BulkCommitResponse(inserted=inserted, updated=updated, failed=len(row_errors))
