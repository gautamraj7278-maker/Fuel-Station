from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime, date
import csv
import io

from app.database import get_db
from app import models, schemas
from app.routers.auth import require_admin

router = APIRouter()


def _interpolate_volume(points: List[models.TankCalibrationPoint], dips_mm: float) -> float:
    if dips_mm < 0:
        raise HTTPException(status_code=400, detail="Dips must be >= 0")

    pts = sorted(points, key=lambda p: float(p.dips_mm))
    if not pts:
        raise HTTPException(status_code=400, detail="No calibration points found for tank")

    if dips_mm < float(pts[0].dips_mm) or dips_mm > float(pts[-1].dips_mm):
        raise HTTPException(
            status_code=400,
            detail=f"Dip {dips_mm}mm is outside calibration range ({pts[0].dips_mm}mm - {pts[-1].dips_mm}mm)",
        )

    for p in pts:
        if float(p.dips_mm) == float(dips_mm):
            return float(p.volume_in_litres)

    lower = None
    upper = None
    for p in pts:
        if float(p.dips_mm) < dips_mm:
            lower = p
        elif float(p.dips_mm) > dips_mm:
            upper = p
            break

    if lower is None or upper is None:
        raise HTTPException(status_code=400, detail="Unable to interpolate dip")

    x0 = float(lower.dips_mm)
    y0 = float(lower.volume_in_litres)
    x1 = float(upper.dips_mm)
    y1 = float(upper.volume_in_litres)

    if x1 == x0:
        return float(y0)

    ratio = (dips_mm - x0) / (x1 - x0)
    return float(y0 + ratio * (y1 - y0))


def _get_calibration_points(db: Session, tank_id: int) -> List[models.TankCalibrationPoint]:
    return (
        db.query(models.TankCalibrationPoint)
        .filter(models.TankCalibrationPoint.tank_id == tank_id)
        .order_by(models.TankCalibrationPoint.dips_mm.asc())
        .all()
    )


def _get_or_auto_opening(db: Session, *, tank_id: int, business_date: date) -> Optional[models.TankDipReading]:
    opening = (
        db.query(models.TankDipReading)
        .filter(
            models.TankDipReading.tank_id == tank_id,
            models.TankDipReading.business_date == business_date,
            models.TankDipReading.dip_type == models.TankDipType.OPENING,
        )
        .first()
    )
    if opening:
        return opening

    prev_date = business_date.fromordinal(business_date.toordinal() - 1)
    prev_closing = (
        db.query(models.TankDipReading)
        .filter(
            models.TankDipReading.tank_id == tank_id,
            models.TankDipReading.business_date == prev_date,
            models.TankDipReading.dip_type == models.TankDipType.CLOSING,
        )
        .first()
    )
    if not prev_closing:
        return None

    auto = models.TankDipReading(
        tank_id=tank_id,
        business_date=business_date,
        dip_type=models.TankDipType.OPENING,
        dips_mm=float(prev_closing.dips_mm),
        computed_volume_litres=prev_closing.computed_volume_litres,
        manual_volume_litres=prev_closing.manual_volume_litres,
        is_auto=True,
        created_by_user_id=None,
    )
    db.add(auto)
    db.flush()
    return auto


def _ensure_buffer_tank(db: Session, product: models.Product) -> models.Tank:
    name = f"BUFFER-{product.product_name}".upper()
    existing = db.query(models.Tank).filter(models.Tank.tank_name == name).first()
    if existing:
        return existing

    buffer_tank = models.Tank(
        tank_name=name,
        product_id=product.id,
        capacity=10**12,  # effectively unlimited virtual tank
        current_volume=0.0,
        is_buffer=True,
        remarks="Auto-created buffer tank",
    )
    db.add(buffer_tank)
    db.flush()
    return buffer_tank


@router.post("/", response_model=schemas.Tank, status_code=status.HTTP_201_CREATED)
def create_tank(
    tank: schemas.TankCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    existing = db.query(models.Tank).filter(models.Tank.tank_name == tank.tank_name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Tank name already exists")

    product = db.query(models.Product).filter(models.Product.id == tank.product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    db_tank = models.Tank(**tank.dict())
    db.add(db_tank)
    db.commit()
    db.refresh(db_tank)
    return db_tank


@router.get("/", response_model=List[schemas.Tank])
def list_tanks(
    product_id: Optional[int] = None,
    is_buffer: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    q = db.query(models.Tank).filter(models.Tank.is_deleted == False)  # noqa: E712
    if product_id is not None:
        q = q.filter(models.Tank.product_id == product_id)
    if is_buffer is not None:
        q = q.filter(models.Tank.is_buffer == is_buffer)
    return q.order_by(models.Tank.tank_name.asc()).all()


@router.get("/deleted", response_model=List[schemas.DeletedItem])
def list_deleted_tanks(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    rows = (
        db.query(models.Tank)
        .filter(models.Tank.is_deleted == True)  # noqa: E712
        .order_by(models.Tank.deleted_at.desc(), models.Tank.id.desc())
        .all()
    )
    return [
        schemas.DeletedItem(
            id=r.id,
            label=r.tank_name,
            deleted_at=r.deleted_at,
            deleted_by_user_id=r.deleted_by_user_id,
            deleted_by_username=(r.deleted_by.username if r.deleted_by else None),
        )
        for r in rows
    ]


@router.post("/deleted/{tank_id}/restore", response_model=schemas.Tank)
def restore_tank(
    tank_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    tank = (
        db.query(models.Tank)
        .filter(models.Tank.id == tank_id, models.Tank.is_deleted == True)  # noqa: E712
        .first()
    )
    if not tank:
        raise HTTPException(status_code=404, detail="Deleted tank not found")

    tank.is_deleted = False
    tank.deleted_at = None
    tank.deleted_by_user_id = None
    db.commit()
    db.refresh(tank)
    return tank


@router.delete("/deleted/{tank_id}", status_code=status.HTTP_204_NO_CONTENT)
def purge_deleted_tank(
    tank_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    tank = (
        db.query(models.Tank)
        .filter(models.Tank.id == tank_id, models.Tank.is_deleted == True)  # noqa: E712
        .first()
    )
    if not tank:
        raise HTTPException(status_code=404, detail="Deleted tank not found")

    try:
        db.delete(tank)
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Cannot purge tank because it is referenced by other records")
    return None


@router.get("/{tank_id}/compute-volume", response_model=schemas.TankComputedVolume)
def compute_volume_from_dips(
    tank_id: int,
    dips_mm: float,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    """Compute volume (litres) for a given tank and dip (mm) using its calibration chart."""
    _ = current_user
    tank = (
        db.query(models.Tank)
        .filter(models.Tank.id == tank_id, models.Tank.is_deleted == False)  # noqa: E712
        .first()
    )
    if not tank:
        raise HTTPException(status_code=404, detail="Tank not found")

    points = _get_calibration_points(db, tank_id)
    volume = _interpolate_volume(points, float(dips_mm))
    return schemas.TankComputedVolume(tank_id=tank_id, dips_mm=float(dips_mm), volume_litres=float(volume))


@router.put("/{tank_id}", response_model=schemas.Tank)
def update_tank(
    tank_id: int,
    payload: schemas.TankUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    tank = (
        db.query(models.Tank)
        .filter(models.Tank.id == tank_id, models.Tank.is_deleted == False)  # noqa: E712
        .first()
    )
    if not tank:
        raise HTTPException(status_code=404, detail="Tank not found")

    update_data = payload.dict(exclude_unset=True)
    if "tank_name" in update_data:
        existing = db.query(models.Tank).filter(models.Tank.tank_name == update_data["tank_name"]).first()
        if existing and existing.id != tank_id:
            raise HTTPException(status_code=400, detail="Tank name already exists")

    if "product_id" in update_data:
        product = db.query(models.Product).filter(models.Product.id == update_data["product_id"]).first()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

    for field, value in update_data.items():
        setattr(tank, field, value)

    db.commit()
    db.refresh(tank)
    return tank


@router.post("/buffer/clear")
def clear_buffer_tanks(
    product_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    _ = current_user
    q = db.query(models.Tank).filter(models.Tank.is_buffer == True)  # noqa: E712
    if product_id is not None:
        q = q.filter(models.Tank.product_id == product_id)

    tanks = q.all()
    if not tanks:
        return {"cleared": 0}

    cleared = 0
    for tank in tanks:
        prev = float(tank.current_volume or 0.0)
        if prev == 0.0:
            continue
        tank.current_volume = 0.0
        db.add(
            models.TankStockLog(
                tank_id=tank.id,
                action="buffer_clear",
                quantity=float(-prev),
                previous_volume=prev,
                new_volume=0.0,
                notes="Manual buffer reset",
                created_by_user_id=current_user.id,
            )
        )
        cleared += 1

    db.commit()
    return {"cleared": cleared}


@router.delete("/{tank_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_tank(
    tank_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    tank = (
        db.query(models.Tank)
        .filter(models.Tank.id == tank_id, models.Tank.is_deleted == False)  # noqa: E712
        .first()
    )
    if not tank:
        raise HTTPException(status_code=404, detail="Tank not found")

    if db.query(models.Nozzle).filter(models.Nozzle.tank_id == tank_id).first():
        raise HTTPException(status_code=400, detail="Cannot delete tank because nozzles reference it")
    if db.query(models.TankTransfer).filter(models.TankTransfer.from_tank_id == tank_id).first() or db.query(models.TankTransfer).filter(models.TankTransfer.to_tank_id == tank_id).first():
        raise HTTPException(status_code=400, detail="Cannot delete tank because transfers reference it")

    tank.is_deleted = True
    tank.deleted_at = datetime.utcnow()
    tank.deleted_by_user_id = current_user.id
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Cannot delete tank because it is referenced by other records")
    return None


@router.post("/{tank_id}/calibration/upload", status_code=status.HTTP_200_OK)
def upload_calibration(
    tank_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    tank = db.query(models.Tank).filter(models.Tank.id == tank_id).first()
    if not tank:
        raise HTTPException(status_code=404, detail="Tank not found")

    filename = (file.filename or "").lower()
    content = file.file.read()

    points: List[models.TankCalibrationPoint] = []

    if filename.endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            dips_mm = float(row.get('dips_mm') or 0)
            volume = float(row.get('volume_in_litres') or 0)
            points.append(models.TankCalibrationPoint(tank_id=tank_id, dips_mm=dips_mm, volume_in_litres=volume))

    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        try:
            import openpyxl  # type: ignore
        except Exception:
            raise HTTPException(
                status_code=400,
                detail="XLSX upload requires 'openpyxl'. Install it in backend environment.",
            )

        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_dips = headers.index('dips_mm') if 'dips_mm' in headers else None
        idx_vol = headers.index('volume_in_litres') if 'volume_in_litres' in headers else None
        if idx_dips is None or idx_vol is None:
            raise HTTPException(status_code=400, detail="Calibration sheet must contain 'dips_mm' and 'volume_in_litres' columns")

        for row in ws.iter_rows(min_row=2):
            dips_cell = row[idx_dips].value
            vol_cell = row[idx_vol].value
            if dips_cell is None or vol_cell is None:
                continue
            points.append(
                models.TankCalibrationPoint(
                    tank_id=tank_id,
                    dips_mm=float(dips_cell),
                    volume_in_litres=float(vol_cell),
                )
            )

    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload .csv or .xlsx")

    if not points:
        raise HTTPException(status_code=400, detail="No calibration points found")

    # Replace existing calibration points
    db.query(models.TankCalibrationPoint).filter(models.TankCalibrationPoint.tank_id == tank_id).delete()
    for p in points:
        db.add(p)

    tank.calibration_date = tank.calibration_date or datetime.utcnow().date()

    db.commit()

    return {"tank_id": tank_id, "points_uploaded": len(points)}


@router.post("/transfer", response_model=schemas.TankTransfer, status_code=status.HTTP_201_CREATED)
def transfer_between_tanks(
    payload: schemas.TankTransferCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    from_tank = db.query(models.Tank).filter(models.Tank.id == payload.from_tank_id).first()
    to_tank = db.query(models.Tank).filter(models.Tank.id == payload.to_tank_id).first()
    if not from_tank or not to_tank:
        raise HTTPException(status_code=404, detail="Tank not found")

    if from_tank.product_id != payload.product_id or to_tank.product_id != payload.product_id:
        raise HTTPException(status_code=400, detail="Tanks must match the transfer product")

    if from_tank.current_volume < payload.volume:
        raise HTTPException(status_code=400, detail="Insufficient volume in source tank")

    if (to_tank.current_volume + payload.volume) > to_tank.capacity:
        raise HTTPException(status_code=400, detail="Destination tank capacity exceeded")

    from_tank.current_volume -= payload.volume
    to_tank.current_volume += payload.volume

    transfer = models.TankTransfer(
        from_tank_id=payload.from_tank_id,
        to_tank_id=payload.to_tank_id,
        product_id=payload.product_id,
        volume=payload.volume,
        transfer_type=models.TankTransferType(payload.transfer_type.value),
        user_id=current_user.id,
    )
    db.add(transfer)
    db.commit()
    db.refresh(transfer)
    return transfer


@router.get("/dips/daily", response_model=List[schemas.TankDipDailyItem])
def get_daily_dips(
    business_date: date,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    tanks = db.query(models.Tank).order_by(models.Tank.tank_name.asc()).all()
    items: List[schemas.TankDipDailyItem] = []
    for t in tanks:
        opening = _get_or_auto_opening(db, tank_id=t.id, business_date=business_date)
        closing = (
            db.query(models.TankDipReading)
            .filter(
                models.TankDipReading.tank_id == t.id,
                models.TankDipReading.business_date == business_date,
                models.TankDipReading.dip_type == models.TankDipType.CLOSING,
            )
            .first()
        )
        items.append(
            schemas.TankDipDailyItem(
                tank_id=t.id,
                tank_name=t.tank_name,
                product_id=t.product_id,
                business_date=business_date,
                opening=opening,
                closing=closing,
            )
        )
    db.commit()  # persist any auto-openings
    return items


@router.post("/dips", response_model=schemas.TankDipReading, status_code=status.HTTP_201_CREATED)
def upsert_dip_reading(
    payload: schemas.TankDipReadingCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    tank = db.query(models.Tank).filter(models.Tank.id == payload.tank_id).first()
    if not tank:
        raise HTTPException(status_code=404, detail="Tank not found")

    existing = (
        db.query(models.TankDipReading)
        .filter(
            models.TankDipReading.tank_id == payload.tank_id,
            models.TankDipReading.business_date == payload.business_date,
            models.TankDipReading.dip_type == models.TankDipType(payload.dip_type.value),
        )
        .first()
    )

    computed = None
    points = _get_calibration_points(db, payload.tank_id)
    try:
        computed = _interpolate_volume(points, float(payload.dips_mm))
    except HTTPException:
        if payload.manual_volume_litres is None:
            raise

    if existing:
        existing.dips_mm = float(payload.dips_mm)
        existing.computed_volume_litres = computed
        existing.manual_volume_litres = float(payload.manual_volume_litres) if payload.manual_volume_litres is not None else None
        existing.is_auto = False
        existing.created_by_user_id = current_user.id
        db.commit()
        db.refresh(existing)
        return existing

    record = models.TankDipReading(
        tank_id=payload.tank_id,
        business_date=payload.business_date,
        dip_type=models.TankDipType(payload.dip_type.value),
        dips_mm=float(payload.dips_mm),
        computed_volume_litres=computed,
        manual_volume_litres=float(payload.manual_volume_litres) if payload.manual_volume_litres is not None else None,
        is_auto=False,
        created_by_user_id=current_user.id,
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.get("/dips/closing-required", response_model=schemas.ClosingRequiredResponse)
def closing_required(
    business_date: date,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    tanks = db.query(models.Tank.id).all()
    tank_ids = [t[0] for t in tanks]
    if not tank_ids:
        return schemas.ClosingRequiredResponse(required=False, business_date=business_date, missing_tank_ids=[])

    missing = []
    for tank_id in tank_ids:
        closing = (
            db.query(models.TankDipReading.id)
            .filter(
                models.TankDipReading.tank_id == tank_id,
                models.TankDipReading.business_date == business_date,
                models.TankDipReading.dip_type == models.TankDipType.CLOSING,
            )
            .first()
        )
        if not closing:
            missing.append(tank_id)

    return schemas.ClosingRequiredResponse(required=len(missing) > 0, business_date=business_date, missing_tank_ids=missing)
